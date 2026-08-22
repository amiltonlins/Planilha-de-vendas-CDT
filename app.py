#!/usr/bin/env python3
"""Painel Comercial Afogados — dashboard, histórico diário e gestão de vendedores."""
from __future__ import annotations
import base64, copy, csv, hmac, html, io, json, os, re, tempfile, time, unicodedata, uuid, zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree as ET

from gerar_painel import ROOT, build_sheets, summarize, tier_value, write_xlsx, month_weeks, month_weeks

NS={"m":"http://schemas.openxmlformats.org/spreadsheetml/2006/main","r":"http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
PUBLISHED_PATH=Path(os.environ.get("PAINEL_DATA_PATH",ROOT/"data"/"dados_publicados.json"))
BASE_SELLER_DEFAULTS=[]
ALIASES={
 "data_venda":("data venda","data da venda","data","dt venda","criado em","data cadastro"),
 "vendedor":("vendedor","colaborador","consultor","nome vendedor","usuario","responsavel"),
 "setor":("setor","equipe","canal","franquia"),
 "categoria":("categoria","tipo vendedor","perfil"),
 "neoenergia":("nome","neoenergia","neoenergia celpe","produto","prospeccao","prospeccao produto","convenio","meio"),
 "adimplencia_m2":("adimplencia m2","m2","pago m2","status m2","adimplencia"),
 "status":("status","situacao","estado"),
 "id_venda":("id venda","id","codigo","numero","proposta","matricula")
}

def normalize_text(value):
    text=unicodedata.normalize("NFKD",str(value or "")).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+"," ",text).strip()

def detect_columns(headers):
    normalized={normalize_text(h):h for h in headers if h is not None}; result={}
    for target,aliases in ALIASES.items():
        for alias in aliases:
            if normalize_text(alias) in normalized:
                result[target]=normalized[normalize_text(alias)]; break
        if target not in result:
            for key,original in normalized.items():
                if any(len(normalize_text(a))>3 and normalize_text(a) in key for a in aliases):
                    result[target]=original; break
    if not {"data_venda","vendedor"}<=result.keys():
        raise ValueError("Não foi possível identificar as colunas de data e vendedor.")
    return result

def excel_date(value):
    if isinstance(value,(int,float)) or (str(value).replace(".","",1).isdigit() and float(value)>1000):
        return (date(1899,12,30)+timedelta(days=int(float(value)))).isoformat()
    text=str(value).strip()
    for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%Y-%m-%d %H:%M:%S","%d/%m/%Y %H:%M"):
        try:return datetime.strptime(text,fmt).date().isoformat()
        except ValueError:pass
    raise ValueError(f"Data inválida: {text}")

def rows_from_csv(data):
    text=data.decode("utf-8-sig",errors="replace")
    try:dialect=csv.Sniffer().sniff(text[:4096],delimiters=",;\t|")
    except csv.Error:dialect=csv.excel
    return list(csv.DictReader(io.StringIO(text),dialect=dialect))

def _xlsx_sheets(data):
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        shared=[]
        if "xl/sharedStrings.xml" in z.namelist():
            root=ET.fromstring(z.read("xl/sharedStrings.xml"))
            shared=["".join(x.text or "" for x in si.iterfind(".//m:t",NS)) for si in root.findall("m:si",NS)]
        workbook=ET.fromstring(z.read("xl/workbook.xml"))
        relroot=ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rels={x.attrib["Id"]:x.attrib["Target"] for x in relroot}
        for sheet in workbook.findall(".//m:sheet",NS):
            target=rels[sheet.attrib[f"{{{NS['r']}}}id"]]
            path=target.lstrip("/"); path=path if path.startswith("xl/") else "xl/"+path
            root=ET.fromstring(z.read(path)); table=[]
            for row in root.findall(".//m:sheetData/m:row",NS):
                values={}
                for cell in row.findall("m:c",NS):
                    letters=re.match(r"[A-Z]+",cell.attrib["r"]).group(); idx=0
                    for ch in letters:idx=idx*26+ord(ch)-64
                    typ=cell.attrib.get("t"); value=""
                    if typ=="inlineStr":value="".join(x.text or "" for x in cell.findall(".//m:t",NS))
                    else:
                        node=cell.find("m:v",NS); value=node.text if node is not None else ""
                        if typ=="s" and value:value=shared[int(value)]
                    values[idx-1]=value
                if values:table.append([values.get(i,"") for i in range(max(values)+1)])
            yield sheet.attrib["name"],table

def rows_from_xlsx(data):
    best=None
    for _,table in _xlsx_sheets(data):
        for index,row in enumerate(table[:30]):
            try:mapping=detect_columns(row)
            except ValueError:continue
            candidate=(len(mapping),row,table[index+1:])
            if best is None or candidate[0]>best[0]:best=candidate
    if not best:raise ValueError("Nenhuma aba de vendas reconhecida no XLSX.")
    _,headers,body=best
    return [{str(headers[i]):row[i] if i<len(row) else "" for i in range(len(headers)) if str(headers[i]).strip()} for row in body if any(str(x).strip() for x in row)]

def canonicalize(raw_rows,base_config):
    if not raw_rows:raise ValueError("O relatório está vazio.")
    mapping=detect_columns(raw_rows[0].keys()); output=[]; seen=set()
    valid={normalize_text(x) for x in base_config["status_validos"]}
    for number,row in enumerate(raw_rows,1):
        try:day=excel_date(row.get(mapping["data_venda"],""))
        except ValueError:continue
        seller=str(row.get(mapping["vendedor"],"")).strip()
        if not seller:continue
        product=str(row.get(mapping.get("neoenergia",""),""))
        status=str(row.get(mapping.get("status",""),"Aprovada") or "Aprovada")
        identifier=str(row.get(mapping.get("id_venda",""),"") or f"UPLOAD-{day}-{seller}-{number}")
        if identifier in seen or normalize_text(status) not in valid:continue
        seen.add(identifier)
        output.append({
            "data_venda":datetime.strptime(day,"%Y-%m-%d").date(),
            "vendedor":seller,
            "setor":str(row.get(mapping.get("setor",""),"NÃO INFORMADO") or "NÃO INFORMADO"),
            "categoria":str(row.get(mapping.get("categoria",""),"Vendedor") or "Vendedor"),
            "neoenergia":"Sim" if "neoenergia celpe" in normalize_text(product) or normalize_text(product) in ("neoenergia","sim","1","true") else "Não",
            "adimplencia_m2":str(row.get(mapping.get("adimplencia_m2",""),"0") or "0"),
            "status":status,"id_venda":identifier})
    if not output:raise ValueError("Nenhuma venda válida após filtros de status e duplicidade.")
    return output,mapping

def merge_registry(base,current):
    # A planilha importada é a fonte de verdade para a existência de vendedores.
    # Não reintroduzir vendedores pré-cadastrados do config base na operação diária.
    # As configurações administrativas já persistidas continuam sendo preservadas.
    cfg=copy.deepcopy(current or base)
    if current is not None:
        cfg["vendedores"]=copy.deepcopy(current.get("vendedores",[]))
    # Migração de compatibilidade: versões anteriores podiam deixar um vendedor
    # ATIVO com categoria "Canal Nacional". Nesse estado ele ficava invisível no
    # ranking porque a elegibilidade exige vendedor ativo + franquia + categoria local.
    # Se o gestor marcou o vendedor para exibir no Dashboard/Ranking, essa intenção
    # administrativa prevalece e o cadastro é normalizado como vendedor da franquia.
    for seller in cfg.get("vendedores",[]):
        if seller.get("ativo",False):
            seller["pertence_franquia"]=True
            if normalize_text(seller.get("categoria")) in {"canal nacional",""}:
                seller["categoria"]="Vendedor"
    return cfg

def prepare_config(base,rows,month,year):
    cfg=copy.deepcopy(base); cfg["mes"],cfg["ano"]=month,year
    existing={normalize_text(x["vendedor"]):x for x in cfg.get("vendedores",[])}; sellers=[]

    # Um mesmo vendedor pode chegar do sistema com diferenças de caixa, acento
    # ou espaços. A configuração precisa ter UMA identidade por nome normalizado.
    grouped={}
    for row in rows:
        key=normalize_text(row.get("vendedor",""))
        if not key:continue
        grouped.setdefault(key,[]).append(row)

    for key in sorted(grouped):
        group=grouped[key]
        sample=group[0]
        previous=existing.get(key,{})
        # Se já existe cadastro, preserva o nome oficial cadastrado; caso contrário
        # usa a grafia mais frequente recebida no relatório.
        if previous.get("vendedor"):
            name=previous["vendedor"]
        else:
            variants={}
            for item in group:
                raw=str(item.get("vendedor","")).strip()
                variants[raw]=variants.get(raw,0)+1
            name=max(variants,key=lambda value:(variants[value],len(value)))
        inferred=next((label for label in ("Website","ADM","Freelance") if normalize_text(label) in key),sample["categoria"])
        registered=bool(previous)
        # Vendedor novo nasce apenas como identificado pela planilha. As decisões
        # administrativas de equipe/visibilidade vêm depois e não filtram suas vendas.
        belongs=previous.get("pertence_franquia",False)
        category=previous.get("categoria",inferred if registered else "Vendedor")
        team_value=normalized_team(previous.get("equipe"),previous or {"setor":sample["setor"],"categoria":category})
        # Compatibilidade com cadastros antigos: quando o estado persistido foi perdido
        # em um redeploy, recuperamos somente vendedores conhecidos no config base.
        # Isso não cria vendas nem altera cálculos; apenas restaura a intenção de exibição.
        fallback_base=next((x for x in BASE_SELLER_DEFAULTS if normalize_text(x.get("vendedor"))==key),{})
        active_default=bool(fallback_base.get("ativo",False)) if not registered else False
        franchise_default=bool(fallback_base.get("pertence_franquia",False)) if not registered else False
        if not registered and fallback_base:
            belongs=franchise_default
            category=fallback_base.get("categoria",category)
            team_value=normalized_team(fallback_base.get("equipe"),fallback_base)
        sellers.append({"vendedor":name,"setor":previous.get("setor",sample["setor"]),"equipe":team_value,"categoria":category,
            "pertence_franquia":belongs,"classificado":previous.get("classificado",registered or bool(fallback_base)),"ativo":previous.get("ativo",active_default if fallback_base else registered),
            "experiencia":previous.get("experiencia",False),"meta_individual":previous.get("meta_individual",70),
            "trabalha_sabado":previous.get("trabalha_sabado",True),"trabalha_domingo":previous.get("trabalha_domingo",False),
            "data_inicio":previous.get("data_inicio",f"{year}-01-01"),"data_desligamento":previous.get("data_desligamento",""),"folgas":previous.get("folgas",[])})
    cfg["vendedores"]=sellers; return cfg

def ensure_import_history(rows,history):
    history=[dict(x) for x in (history or [])]
    for i,h in enumerate(history):
        if not h.get("importacao_id"):
            seed=f'{h.get("data_importacao","")}|{h.get("arquivo","")}|{i}'
            h["importacao_id"]="legacy-"+str(uuid.uuid5(uuid.NAMESPACE_URL,seed))
        h.setdefault("status","Ativa")
        h.setdefault("vendedores",0)
        h.setdefault("usuario","")
    latest_by_day={}
    for i,h in enumerate(history):
        if h.get("status")=="Excluída":continue
        for day in h.get("dias",[]):latest_by_day[day]=i
    for i,h in enumerate(history):
        if h.get("status")=="Excluída":continue
        days=h.get("dias",[])
        h["status"]="Ativa" if any(latest_by_day.get(day)==i for day in days) else "Substituída"
    active_by_day={}
    for h in history:
        if h.get("status")!="Ativa":continue
        for day in h.get("dias",[]):active_by_day[day]=h["importacao_id"]
    for row in rows:
        if not row.get("importacao_id"):
            row["importacao_id"]=active_by_day.get(row["data_venda"].isoformat(),"")
    return rows,history

def mark_replaced_imports(history,imported_days):
    day_keys={d.isoformat() if hasattr(d,"isoformat") else str(d) for d in imported_days}
    for h in history:
        if h.get("status","Ativa")=="Ativa" and day_keys.intersection(set(h.get("dias",[]))):
            h["status"]="Substituída"
    return history

def merge_daily_history(current_rows,incoming_rows):
    if not incoming_rows:return current_rows,[]
    imported_days={x["data_venda"] for x in incoming_rows}
    kept=[x for x in current_rows if x["data_venda"] not in imported_days]
    dedup={}
    for row in kept+incoming_rows:dedup[str(row["id_venda"])]=row
    merged=sorted(dedup.values(),key=lambda x:(x["data_venda"],normalize_text(x["vendedor"]),str(x["id_venda"])))
    return merged,sorted(imported_days)

def serialize_row(row):
    out=dict(row); out["data_venda"]=row["data_venda"].isoformat(); return out

def save_published(rows,cfg,source_name,history=None,updated_at=None):
    updated_at=updated_at or datetime.now(); PUBLISHED_PATH.parent.mkdir(parents=True,exist_ok=True)
    payload={"atualizado_em":updated_at.isoformat(timespec="seconds"),"arquivo":Path(source_name).name,
             "config":cfg,"historico_importacoes":history or [],"vendas":[serialize_row(row) for row in rows]}
    temporary=PUBLISHED_PATH.with_suffix(".tmp"); temporary.write_text(json.dumps(payload,ensure_ascii=False),encoding="utf-8"); temporary.replace(PUBLISHED_PATH)

def load_published(base):
    if PUBLISHED_PATH.exists():
        payload=json.loads(PUBLISHED_PATH.read_text(encoding="utf-8")); rows=payload.get("vendas",[])
        for row in rows:row["data_venda"]=datetime.strptime(row["data_venda"],"%Y-%m-%d").date()
        rows,history=ensure_import_history(rows,payload.get("historico_importacoes",[]))
        payload["historico_importacoes"]=history
        return rows,merge_registry(base,payload.get("config",base)),payload
    raw=rows_from_csv((ROOT/"dados_exemplo.csv").read_bytes()); rows,_=canonicalize(raw,base); cfg=prepare_config(base,rows,base["mes"],base["ano"])
    return rows,cfg,{"atualizado_em":datetime.now().isoformat(timespec="seconds"),"arquivo":"dados_exemplo.csv","demonstracao":True,"historico_importacoes":[]}

TEAM_OPTIONS=("Equipe Interna","Equipe Externa","Outros Canais")

def normalized_team(value, seller=None):
    raw=normalize_text(value)
    canonical={normalize_text(x):x for x in TEAM_OPTIONS}
    if raw in canonical:return canonical[raw]
    legacy=normalize_text((seller or {}).get("setor",value))
    category=normalize_text((seller or {}).get("categoria",""))
    combined=f"{legacy} {category}"
    if any(token in combined for token in ("pap","extern","porta a porta")):return "Equipe Externa"
    if any(token in combined for token in ("adm","website","freelance","canal nacional")):return "Outros Canais"
    return "Equipe Interna"

def apply_team_labels(summary,cfg):
    registry={normalize_text(x.get("vendedor")):x for x in cfg.get("vendedores",[])}
    for item in summary:
        seller=registry.get(normalize_text(item.get("vendedor")),{})
        item["equipe"]=normalized_team(seller.get("equipe"),seller)
    return summary

def team_sales_totals(summary,cfg):
    """Soma vendas por equipe usando a configuração gerencial, inclusive vendedores ocultos."""
    registry={normalize_text(x.get("vendedor")):x for x in cfg.get("vendedores",[])}
    totals={"Equipe Interna":0,"Equipe Externa":0}
    excluded={"website","adm","freelance","canal nacional"}
    for item in summary:
        seller=registry.get(normalize_text(item.get("vendedor")),{})
        if not seller.get("pertence_franquia",False):
            continue
        if normalize_text(seller.get("categoria",item.get("categoria",""))) in excluded:
            continue
        team=normalized_team(seller.get("equipe"),seller)
        if team in totals:
            totals[team]+=int(item.get("vendas",0) or 0)
    return totals

def performance(media):
    media=float(media or 0)
    if media>=2.0:return "Azul","#0891B2","cyan"
    if media>=1.5:return "Verde","#16A34A","green"
    if media>=1.0:return "Amarelo","#F59E0B","yellow"
    return "Vermelho","#DC2626","red"

def money(value):return f"R$ {value:,.2f}".replace(",","X").replace(".",",").replace("X",".")
def pct(value):return f"{value:.1%}".replace(".",",")
FIXED_DASHBOARD_USERS={
    "Amilton Lins":("amilton lins","hamilton lins"),
    "Sheyla Santos":("sheyla santos","sheila santos"),
    "Joice Larissa":("joice larissa","joyce larissa"),
    "Rafael Salgado":("rafael salgado","raphael salgado"),
}

def first_name(value):
    return str(value or "").strip().split()[0] if str(value or "").strip() else ""

def first_two_names(value):
    parts=str(value or "").strip().split()
    return " ".join(parts[:2]) if len(parts)>=2 else ""

def authorized_dashboard_users(cfg):
    identities={}
    for display,aliases in FIXED_DASHBOARD_USERS.items():
        for alias in aliases:
            identities[normalize_text(alias)]={"display":display,"full":display,"fixed":True}
    for seller in cfg.get("vendedores",[]):
        if not seller.get("ativo",False):continue
        full=str(seller.get("vendedor","")).strip()
        key=normalize_text(first_two_names(full))
        if not key:continue
        identities.setdefault(key,{"display":first_two_names(full),"full":full,"fixed":False})
    return identities

def authenticate_dashboard_name(cfg,provided_first,provided_full=""):
    candidate=(provided_full or provided_first or "").strip()
    key=normalize_text(first_two_names(candidate))
    if not key:return None,"invalid"
    selected=authorized_dashboard_users(cfg).get(key)
    return (selected["display"],"ok") if selected else (None,"invalid")

def is_mobile_client(st):
    try:
        ua=str(st.context.headers.get("User-Agent","")).lower()
    except Exception:
        return False
    return any(token in ua for token in ("iphone","ipad","ipod","android","mobile"))

def auth_signing_key(st):
    try:
        key=str(st.secrets.get("DASHBOARD_AUTH_KEY","") or st.secrets.get("GESTOR_SENHA",""))
    except Exception:
        key=os.environ.get("DASHBOARD_AUTH_KEY","") or os.environ.get("GESTOR_SENHA","")
    return key

def issue_dashboard_token(st,user):
    key=auth_signing_key(st)
    if not key:return ""
    expiry=int(time.time())+12*60*60
    payload=f"{user}|{expiry}"
    sig=hmac.new(key.encode("utf-8"),payload.encode("utf-8"),"sha256").hexdigest()
    raw=f"{payload}|{sig}".encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")

def validate_dashboard_token(st,cfg,token):
    key=auth_signing_key(st)
    if not key or not token:return None
    try:
        padded=str(token)+"="*((4-len(str(token))%4)%4)
        raw=base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        user,expiry_text,sig=raw.rsplit("|",2)
        if int(expiry_text)<int(time.time()):return None
        payload=f"{user}|{expiry_text}"
        expected=hmac.new(key.encode("utf-8"),payload.encode("utf-8"),"sha256").hexdigest()
        if not hmac.compare_digest(sig,expected):return None
        allowed={normalize_text(x["display"]) for x in authorized_dashboard_users(cfg).values()}
        return user if normalize_text(user) in allowed else None
    except Exception:
        return None

def validate_xlsx_bytes(data,required_sheet=None):
    if not data:raise ValueError("Arquivo XLSX vazio.")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            if archive.testzip() is not None:raise ValueError("Estrutura ZIP do XLSX inválida.")
            names=set(archive.namelist())
            for required in ("[Content_Types].xml","xl/workbook.xml","xl/_rels/workbook.xml.rels"):
                if required not in names:raise ValueError(f"Componente XLSX ausente: {required}")
            ET.fromstring(archive.read("xl/workbook.xml"))
        sheets=list(_xlsx_sheets(data))
        if not sheets:raise ValueError("O XLSX não contém planilhas legíveis.")
        if required_sheet and required_sheet not in {name for name,_ in sheets}:raise ValueError(f"A aba {required_sheet} não foi encontrada no XLSX.")
    except (zipfile.BadZipFile,ET.ParseError,KeyError,IndexError,ValueError) as exc:
        if isinstance(exc,ValueError):raise
        raise ValueError(f"XLSX inválido: {exc}") from exc
    return True

def general_report_display(team):
    display=[]
    for item in sorted(team,key=lambda z:(z["vendas"],z["projecao"]),reverse=True):
        display.append(item|{
            "media":f'{item["media"]:.2f}',
            "meta_pct":pct(item["projecao"]/item["meta_individual"] if item["meta_individual"] else 0),
            "neo_pct_fmt":pct(item["neo_pct"]),
            "base_fmt":money(item["base"]),
            "proj_fmt":money(item["comissao_proj"]),
            "neo_proj_fmt":money(item["bonus_neo_proj"]),
            "adim_proj_fmt":money(item["bonus_adim_proj"]),
            "premio_fmt":money(item["premio_total"]),
            "total_proj_fmt":money(item["total_variavel_proj"])})
    return display

def general_report_xlsx_bytes(team,all_days):
    """Gera XLSX padrão e validável pelo Microsoft Excel."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    output=io.BytesIO()
    wb=Workbook()
    ws=wb.active
    ws.title="RELATORIO GERAL"
    headers=["EQUIPE","VENDEDOR","TOTAL","PROJEÇÃO","MÉDIA","ZEROS","% META","NEO","% NEO","PREMIAÇÃO ATUAL","PREMIAÇÃO PROJETADA","BÔNUS NEO PROJ.","BÔNUS (SE) 100% ADIM","SEMANAIS","TOTAL VAR. PROJ."]+[str(d.day) for d in all_days]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill=PatternFill("solid",fgColor="0F172A")
        cell.font=Font(color="FFFFFF",bold=True)
        cell.alignment=Alignment(horizontal="center")
    for item in sorted(team,key=lambda z:(z["vendas"],z["projecao"]),reverse=True):
        meta_pct=item["projecao"]/item["meta_individual"] if item.get("meta_individual") else 0
        values=[item.get("equipe",""),item.get("vendedor",""),int(item.get("vendas",0) or 0),int(item.get("projecao",0) or 0),float(item.get("media",0) or 0),int(item.get("zeros",0) or 0),meta_pct,int(item.get("neo",0) or 0),float(item.get("neo_pct",0) or 0),float(item.get("base",0) or 0),float(item.get("comissao_proj",0) or 0),float(item.get("bonus_neo_proj",0) or 0),float(item.get("bonus_adim_proj",0) or 0),float(item.get("premio_total",0) or 0),float(item.get("total_variavel_proj",0) or 0)]
        elapsed_days=item.get("dias_decorridos",set())
        daily=item.get("diario",{})
        for d in all_days:
            values.append(int(daily.get(d.day,0) or 0) if d.day in elapsed_days else None)
        ws.append(values)
    ws.freeze_panes="A2"
    ws.auto_filter.ref=ws.dimensions
    ws.column_dimensions["A"].width=18
    ws.column_dimensions["B"].width=34
    for col in range(3,16):ws.column_dimensions[get_column_letter(col)].width=18
    for col in range(16,16+len(all_days)):ws.column_dimensions[get_column_letter(col)].width=5
    for row in ws.iter_rows(min_row=2):
        row[6].number_format='0.0%'
        row[8].number_format='0.0%'
        for idx in range(9,15):row[idx].number_format='R$ #,##0.00'
    wb.save(output)
    data=output.getvalue()
    load_workbook(io.BytesIO(data),read_only=True,data_only=True).close()
    return data

def render_general_report(st,team,rows,cfg,summary,all_days,elapsed,official,color):
    st.markdown('<div class="section">Relatório geral da equipe</div>',unsafe_allow_html=True)
    cols=[("equipe","EQUIPE"),("vendedor","VENDEDOR"),("vendas","TOTAL"),("projecao","PROJEÇÃO"),("media","MÉDIA"),("zeros","ZEROS"),("meta_pct","% META"),("neo","NEO"),("neo_pct_fmt","% NEO"),("base_fmt","PREMIAÇÃO ATUAL"),("proj_fmt","PREMIAÇÃO PROJETADA"),("neo_proj_fmt","BÔNUS NEO PROJ."),("adim_proj_fmt","BÔNUS (SE) 100% ADIM"),("premio_fmt","SEMANAIS"),("total_proj_fmt","TOTAL VAR. PROJ.")]+[(d.day,str(d.day)) for d in all_days]
    display=general_report_display(team)
    st.markdown(table_html(display,cols,color,True),unsafe_allow_html=True)

def manager_password(st):
    try:return str(st.secrets["GESTOR_SENHA"])
    except (KeyError,FileNotFoundError):return os.environ.get("GESTOR_SENHA","")
def card(label,value,tone="cyan",sub=""):return f'<div class="metric {tone}"><span>{html.escape(label)}</span><strong>{html.escape(str(value))}</strong><small>{html.escape(sub)}</small></div>'
def cards(st,items,columns=4):
    cols=st.columns(columns)
    for i,item in enumerate(items):cols[i%columns].markdown(card(*item),unsafe_allow_html=True)
def regular(summary):return [x for x in summary if x.get("elegivel_individual",True)]
def channel_name(item):
    if not item.get("pertence_franquia",True) or normalize_text(item.get("categoria"))=="canal nacional":return "CANAL NACIONAL"
    category=normalize_text(item.get("categoria"))
    if category in {"website","adm","freelance"}:return category.upper()
    return "VENDEDORES FRANQUIA"

def table_html(records,columns,row_color=None,daily=False):
    heads="".join(f"<th>{html.escape(str(label))}</th>" for _,label in columns); body=[]
    for rec in records:
        color=row_color(rec) if row_color else "#fff"; cells=[]
        for key,_ in columns:
            value=rec.get(key,""); style=""
            if key=="vendedor" and row_color:style=f"background:{color};color:white;font-weight:900"
            if daily and isinstance(key,int):
                elapsed=rec.get("dias_decorridos",set()); scheduled=rec.get("dias_agendados",set()); diario=rec.get("diario",{})
                if key not in elapsed:style="background:#F1F5F9;color:#94A3B8"; value=""
                elif key in scheduled and diario.get(key,0)==0:style="background:#FEE2E2;color:#B91C1C;font-weight:800"; value=0
                elif diario.get(key,0)>=3:style="background:#CFFAFE;color:#155E75;font-weight:800"; value=diario.get(key,0)
                else:value=diario.get(key,0)
            cells.append(f'<td style="{style}">{html.escape(str(value))}</td>')
        body.append(f'<tr>{"".join(cells)}</tr>')
    return f'<div class="table-wrap"><table class="report"><thead><tr>{heads}</tr></thead><tbody>{"".join(body)}</tbody></table></div>'

def executive_kpis_html(cfg,total,projection,neo,team):
    meta=cfg["meta_empresa"]
    ating=total/meta if meta else 0
    faltam=max(0,meta-total)
    neo_pct=neo/total if total else 0
    zeros=sum(x["zeros"] for x in team)
    return (
        '<div class="exec-grid">'
        f'<div class="exec-card hero"><small>VENDAS REALIZADAS</small><strong>{total}</strong><span>Resultado atual</span></div>'
        f'<div class="exec-card"><small>META DO MÊS</small><strong>{meta}</strong><span>Objetivo comercial</span></div>'
        f'<div class="exec-card projection"><small>PROJEÇÃO</small><strong>{projection}</strong><span>Fechamento estimado</span></div>'
        f'<div class="exec-card attainment"><small>% DA META</small><strong>{pct(ating)}</strong><span>Atingimento atual</span></div>'
        f'<div class="exec-card secondary"><small>FALTAM PARA META</small><strong>{faltam}</strong><span>Vendas necessárias</span></div>'
        f'<div class="exec-card secondary"><small>VENDAS NEO</small><strong>{neo}</strong><span>{pct(neo_pct)} do total</span></div>'
        f'<div class="exec-card secondary"><small>% NEO</small><strong>{pct(neo_pct)}</strong><span>Participação</span></div>'
        '</div>'
    )

def performance_summary_html(counts):
    items=[
        ("Azul",counts.get("Azul",0),"#0891B2"),
        ("Verde",counts.get("Verde",0),"#16A34A"),
        ("Amarelo",counts.get("Amarelo",0),"#F59E0B"),
        ("Vermelho",counts.get("Vermelho",0),"#DC2626"),
    ]
    chips=''.join(
        f'<div class="perf-chip" style="--chip:{color}"><span>{label}</span><strong>{value}</strong></div>'
        for label,value,color in items
    )
    return f'<div class="perf-summary">{chips}</div>'

def channel_summary_html(channels,total):
    items=(("VENDEDORES FRANQUIA",channels.get("VENDEDORES FRANQUIA",0),"blue"),("CANAL NACIONAL",channels.get("CANAL NACIONAL",0),"green"))
    cards=[]
    for name,value,tone in items:
        share=value/total if total else 0
        cards.append(f'<div class="pc-channel-card {tone}"><div class="pc-channel-content"><span>{name}</span><strong>{value}</strong><small>{pct(share)} do total</small><div class="pc-channel-track"><i style="width:{min(100,max(0,share*100)):.1f}%"></i></div></div></div>')
    return '<div class="pc-channel-row">'+''.join(cards)+'</div>'

def team_performance_metrics(summary,team_name,goal):
    local=[]
    for item in summary:
        category=normalize_text(item.get("categoria",""))
        if item.get("equipe")==team_name and item.get("pertence_franquia",True) and category not in {"website","adm","freelance","canal nacional"}:
            local.append(item)
    sales=sum(int(x.get("vendas",0) or 0) for x in local)
    projection=sum(int(x.get("projecao",0) or 0) for x in local)
    elapsed=max((int(x.get("dias",0) or 0) for x in local),default=0)
    planned=max((int(x.get("dias_previstos",0) or 0) for x in local),default=0)
    average=sales/elapsed if elapsed else 0
    missing=max(0,int(goal or 0)-sales)
    remaining=max(0,planned-elapsed)
    needed=missing/remaining if remaining else 0
    attainment=sales/int(goal or 0) if int(goal or 0)>0 else 0
    return {"sales":sales,"projection":projection,"elapsed":elapsed,"planned":planned,"average":average,"missing":missing,"remaining":remaining,"needed":needed,"attainment":attainment}

def team_performance_card_html(title,goal,metrics,tone="internal"):
    width=min(100,max(0,metrics["attainment"]*100))
    return f'''<div class="pc-team-card {tone}"><div class="pc-team-head"><div class="pc-team-title"><b>{html.escape(title.upper())}</b></div><div class="pc-team-goal"><span>META</span><b>{int(goal or 0)}</b></div></div><div class="pc-team-main"><div class="pc-team-sales"><strong>{metrics["sales"]}</strong><span>VENDAS</span></div><div class="pc-team-progress"><b>{pct(metrics["attainment"])}</b><span>DA META</span><div class="pc-team-track"><i style="width:{width:.1f}%"></i></div></div></div><div class="pc-team-stats"><div><span>MÉDIA/DIA</span><strong>{metrics["average"]:.1f}</strong></div><div><span>PROJEÇÃO</span><strong>{metrics["projection"]}</strong></div><div><span>FALTAM</span><strong>{metrics["missing"]}</strong></div><div><span>NECESSÁRIO/DIA</span><strong>{metrics["needed"]:.1f}</strong></div></div></div>'''

def production_channel_dashboard_html(channels,total,summary,cfg,data_until,updated):
    internal_goal=int(cfg.get("meta_equipe_interna",0) or 0); external_goal=int(cfg.get("meta_equipe_externa",0) or 0)
    internal=team_performance_metrics(summary,"Equipe Interna",internal_goal); external=team_performance_metrics(summary,"Equipe Externa",external_goal)
    return f'''<div class="pc-dashboard">{channel_summary_html(channels,total)}<div class="pc-team-grid">{team_performance_card_html("Equipe Interna",internal_goal,internal,"internal")}{team_performance_card_html("Equipe Externa",external_goal,external,"external")}</div></div>'''


def weekly_rank_html(team,week_index):
    ranked=[]
    for x in team:
        sales=x["semanas"][week_index] if week_index < len(x.get("semanas",[])) else 0
        award=x["premios"][week_index] if week_index < len(x.get("premios",[])) else 0
        ranked.append((x,sales,award))
    ranked.sort(key=lambda item:(item[1],item[2]),reverse=True)
    cards=[]
    for pos,(x,sales,award) in enumerate(ranked,1):
        cards.append(
            f'<div class="weekly-rank-card"><div class="weekly-rank-pos">{pos}º</div>'
            f'<div class="weekly-rank-name">{html.escape(str(x["vendedor"]))}</div>'
            f'<div class="weekly-rank-sales"><strong>{sales}</strong><span>VENDAS</span></div>'
            f'<div class="weekly-rank-award"><strong>{money(award)}</strong><span>SEMANAL</span></div></div>'
        )
    return '<div class="weekly-rank-list">'+''.join(cards)+'</div>'

def weekly_seller_history_html(x):
    rows=[]
    for i,sales in enumerate(x.get("semanas",[])):
        award=x["premios"][i] if i < len(x.get("premios",[])) else 0
        rows.append(f'<div class="weekly-history-row"><span>SEMANA {i+1}</span><strong>{sales} vendas</strong><b>{money(award)}</b></div>')
    return (f'<div class="weekly-history"><div class="weekly-history-name">{html.escape(str(x["vendedor"]))}</div>'
            +''.join(rows)+f'<div class="weekly-history-total"><span>TOTAL NO MÊS</span><strong>{x.get("vendas",0)} vendas</strong><b>{money(x.get("premio_total",0))}</b></div></div>')

def weekly_desktop_table(team,max_weeks):
    rows=[]
    for x in sorted(team,key=lambda z:(sum(z.get("semanas",[])),z.get("premio_total",0)),reverse=True):
        cells=[f'<td class="weekly-name">{html.escape(str(x["vendedor"]))}</td>']
        for i in range(max_weeks):
            sales=x["semanas"][i] if i < len(x.get("semanas",[])) else 0
            award=x["premios"][i] if i < len(x.get("premios",[])) else 0
            cells.append(f'<td><strong>{sales}</strong><small>vendas</small></td><td><strong>{money(award)}</strong><small>valor</small></td>')
        cells.append(f'<td><strong>{money(x.get("premio_total",0))}</strong></td>')
        rows.append('<tr>'+''.join(cells)+'</tr>')
    heads=['VENDEDOR']
    for i in range(max_weeks):heads.extend([f'S{i+1} VENDAS',f'S{i+1} VALOR'])
    heads.append('SEMANAIS ACUMULADOS')
    return '<div class="weekly-desktop-table"><table><thead><tr>'+''.join(f'<th>{h}</th>' for h in heads)+'</tr></thead><tbody>'+''.join(rows)+'</tbody></table></div>'

def projection_status_visual(projected_ratio, meta_value=None, projection_value=None):
    try:
        if meta_value in (None, 0, "") or projection_value in (None, ""):
            return "🙂", "AGUARDANDO DADOS"
        ratio=float(projected_ratio)
        if ratio != ratio or ratio < 0:  # NaN ou inválido
            return "🙂", "AGUARDANDO DADOS"
    except (TypeError, ValueError, ZeroDivisionError):
        return "🙂", "AGUARDANDO DADOS"

    if ratio < 0.40:return "😭", "MUITO ABAIXO"
    if ratio < 0.60:return "😟", "ATENÇÃO"
    if ratio < 0.80:return "😐", "PRECISA ACELERAR"
    if ratio < 0.90:return "🙂", "BOM RITMO"
    if ratio < 1.00:return "😄", "QUASE LÁ!"
    if ratio < 1.20:return "😎", "META NO CAMINHO!"
    return "🤩", "VOANDO!"


def projection_status_card(emoji,message):
    return (
        '<div class="seller-kpi projection-status">'
        f'<div class="projection-status-emoji">{emoji}</div>'
        f'<div class="projection-status-message">{html.escape(str(message))}</div>'
        '</div>'
    )

def seller_kpi_card(label,value,sub,cls):
    return f'<div class="seller-kpi {cls}"><small>{html.escape(str(label))}</small><strong>{html.escape(str(value))}</strong><span>{html.escape(str(sub))}</span></div>'

def seller_kpis_html(x):
    meta_pct=x["projecao"]/x["meta_individual"] if x["meta_individual"] else 0
    status,_,_=performance(x["media"])
    commercial=[
        ("VENDAS",x["vendas"],"Produção","primary mobile-duplicate"),
        ("PROJEÇÃO",x["projecao"],f'Meta {x["meta_individual"]}',"level2 mobile-duplicate"),
        ("MÉDIA/DIA",f'{x["media"]:.2f}',f'{x["dias"]} dias',"level2"),
        ("% META",pct(meta_pct),status,"level2"),
        ("ZEROS",x["zeros"],f'Semana {x["zeros_semana"]}',"level3"),
        ("NEO",x["neo"],"Neoenergia","level3"),
        ("% NEO",pct(x["neo_pct"]),"Participação","level3"),
    ]
    awards=[
        ("PREMIAÇÃO ATUAL",money(x["base"]),"Já acumulada","primary mobile-duplicate"),
        ("PREMIAÇÃO PROJETADA",money(x["comissao_proj"]),"Base projetada","level2"),
        ("BÔNUS NEO PROJETADO",money(x["bonus_neo_proj"]),"Projeção","level3"),
        ("BÔNUS (SE) 100% ADIM",money(x["bonus_adim_proj"]),"Condicional","level3"),
        ("SEMANAIS",money(x["premio_total"]),"Acumulado semanal","level3"),
        ("TOTAL VARIÁVEL PROJETADO",money(x["total_variavel_proj"]),"Fechamento estimado","primary total mobile-duplicate"),
    ]
    commercial_html=''.join(seller_kpi_card(*item) for item in commercial)
    awards_html=''.join(seller_kpi_card(*item) for item in awards)
    mobile_primary=[
        ("VENDAS",x["vendas"],"Produção","primary"),
        ("PREMIAÇÃO ATUAL",money(x["base"]),"Já acumulada","primary"),
        ("PROJEÇÃO",x["projecao"],f'Meta {x["meta_individual"]}',"primary mobile-projection"),
        ("TOTAL VARIÁVEL PROJETADO",money(x["total_variavel_proj"]),"Fechamento estimado","primary total"),
    ]
    mobile_html=''.join(seller_kpi_card(*item) for item in mobile_primary)
    return (
        f'<div class="seller-mobile-primary">{mobile_html}</div>'
        '<div class="seller-groups">'
        '<div class="seller-group-title">DESEMPENHO COMERCIAL</div>'
        f'<div class="seller-kpi-grid">{commercial_html}</div>'
        '<div class="seller-group-title award">PREMIAÇÃO</div>'
        f'<div class="seller-kpi-grid award-grid">{awards_html}</div>'
        '</div>'
    )

def ranking_html(ranking,auth_token=""):
    medals=("🥇","🥈","🥉")
    rows=[]
    for i,x in enumerate(ranking):
        _,color,_=performance(x["media"])
        medal=medals[i] if i<3 else f"{i+1}º"
        meta_pct=x["projecao"]/x["meta_individual"] if x["meta_individual"] else 0
        status_emoji,status_message=projection_status_visual(meta_pct,x.get("meta_individual"),x.get("projecao"))
        rows.append(
            f'<div class="rank-row"><div class="rank-pos">{medal}</div>'
            f'<a class="rank-click" href="?{("auth="+html.escape(str(auth_token),quote=True)+"&") if auth_token else ""}seller={html.escape(str(x["vendedor"]),quote=True)}" target="_self">'
            f'<div class="rank-name" style="background:{color}">' 
            f'<div class="rank-seller"><div class="rank-mobile-head"><div><b>{html.escape(x["vendedor"])}</b><small>{html.escape(x.get("equipe","Equipe Interna"))}</small></div><div class="rank-mobile-status"><strong>{status_emoji}</strong><small>{html.escape(status_message)}</small></div></div></div>' 
            f'<div class="rank-inside">'
            f'<span class="main-kpi"><strong>{x["vendas"]}</strong><small>VENDAS</small></span>'
            f'<span class="main-kpi"><strong>{x["projecao"]}</strong><small>PROJEÇÃO</small></span>'
            f'<span><strong>{x["media"]:.2f}</strong><small>MÉDIA/DIA</small></span>'
            f'<span><strong>{x["zeros"]}</strong><small>ZEROS</small></span>'
            f'<span><strong>{pct(meta_pct)}</strong><small>% META</small></span>'
            f'<span class="neo-highlight"><strong>{x["neo"]}</strong><small>NEO</small></span>'
            f'<span class="neo-highlight"><strong>{pct(x["neo_pct"])}</strong><small>% NEO</small></span>'
            f'<span><strong>{money(x["base"])}</strong><small>PREMIAÇÃO ATUAL</small></span>'
            f'<span><strong>{money(x["comissao_proj"])}</strong><small>PREMIAÇÃO PROJ.</small></span>'
            f'<span><strong>{money(x["bonus_neo_proj"])}</strong><small>BÔNUS NEO PROJ.</small></span>'
            f'<span><strong>{money(x["bonus_adim_proj"])}</strong><small>BÔNUS (SE) 100% ADIM</small></span>'
            f'<span><strong>{money(x["premio_total"])}</strong><small>SEMANAIS</small></span>'
            f'<span class="total-highlight"><strong>{money(x["total_variavel_proj"])}</strong><small>TOTAL VAR. PROJ.</small></span>'
            f'<span class="rank-projection-status desktop-status"><strong class="rank-status-emoji">{status_emoji}</strong><small>{html.escape(status_message)}</small></span>'
            f'</div></div></a></div>'
        )
    return '<div class="rank-card">'+''.join(rows)+'</div>' if rows else '<div class="empty-bi">Nenhum vendedor local ativo para exibir no ranking.</div>'

def daily_series(rows,cfg,data_until):
    relevant=[x for x in rows if x["data_venda"].year==cfg["ano"] and x["data_venda"].month==cfg["mes"] and x["data_venda"]<=data_until]
    by_day={}
    for row in relevant:by_day[row["data_venda"].day]=by_day.get(row["data_venda"].day,0)+1
    last=max(1,data_until.day); cumulative=[]; ideal=[]; running=0
    for day in range(1,last+1):
        running+=by_day.get(day,0); cumulative.append(running); ideal.append(round(cfg["meta_empresa"]*day/31))
    return cumulative,ideal

CSS="""<style>
#MainMenu,footer,header,[data-testid="stDecoration"]{display:none!important}
:root{--navy:#0F172A;--ink:#111827;--muted:#64748B;--line:#E2E8F0;--surface:#FFFFFF;--bg:#F4F7FB;--cyan:#0EA5E9;--green:#22C55E;--amber:#F59E0B;--red:#EF4444}
.stApp{background:var(--bg);color:var(--ink)}.block-container{padding:1rem 1.5rem 2.5rem;max-width:1920px}[data-testid="stSidebar"]{display:none!important}
.bi-topbar{background:linear-gradient(110deg,#0F172A,#172554);color:white;border-radius:14px;padding:18px 22px;margin:0 0 10px;box-shadow:0 8px 30px rgba(15,23,42,.12)}.bi-topbar h1{font-size:1.42rem;margin:0;font-weight:800;letter-spacing:-.02em}.bi-topbar p{margin:4px 0 0;color:#CBD5E1;font-size:.78rem}
.section{font-size:.83rem;font-weight:800;color:#334155;padding:5px 0;margin:14px 0 7px;letter-spacing:.035em;text-transform:uppercase}.metric{position:relative;background:var(--surface);border:1px solid var(--line);border-radius:12px;padding:13px 14px;min-height:92px;margin:2px 0 5px;box-shadow:0 2px 10px rgba(15,23,42,.045);overflow:hidden}.metric:before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--cyan)}.metric span{display:block;font-size:.66rem;font-weight:800;color:#64748B;letter-spacing:.05em;text-transform:uppercase}.metric strong{display:block;font-size:1.65rem;line-height:1.15;color:#0F172A;margin-top:8px;font-weight:800}.metric small{display:block;font-size:.68rem;color:#94A3B8;margin-top:5px}.metric.green:before{background:var(--green)}.metric.yellow:before{background:var(--amber)}.metric.red:before{background:var(--red)}
.bi-panel{background:white;border:1px solid var(--line);border-radius:14px;padding:14px 16px;box-shadow:0 2px 10px rgba(15,23,42,.04)}.rank-card{background:white;border:1px solid var(--line);border-radius:14px;overflow:hidden;box-shadow:0 2px 10px rgba(15,23,42,.04)}.rank-row{display:grid;grid-template-columns:54px 1fr;align-items:center;gap:10px;padding:9px 13px;border-bottom:1px solid #EEF2F7}.rank-row:last-child{border-bottom:0}.rank-pos{font-weight:900;text-align:center}.rank-name{display:grid;grid-template-columns:minmax(220px,.9fr) 3.6fr;align-items:start;gap:14px;min-width:0;padding:12px;border-radius:10px;color:white;box-shadow:0 2px 7px rgba(15,23,42,.10)}.rank-seller{display:flex;flex-direction:column;min-width:0;padding-top:4px}.rank-seller b{white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:.94rem}.rank-seller small{font-size:.65rem;color:rgba(255,255,255,.82)}.rank-inside{display:grid;grid-template-columns:repeat(7,minmax(92px,1fr));gap:7px;align-items:stretch}.rank-inside span{display:flex;flex-direction:column;justify-content:center;align-items:center;min-height:54px;padding:5px 7px;border-left:1px solid rgba(255,255,255,.25);text-align:center}.rank-inside strong{font-size:.92rem;line-height:1.08;font-weight:900;white-space:nowrap}.rank-inside small{font-size:.54rem;color:rgba(255,255,255,.84);margin-top:5px}.rank-inside .main-kpi strong{font-size:1.35rem}.rank-inside .neo-highlight{background:rgba(255,255,255,.14);border-radius:8px}.rank-inside .neo-highlight strong{font-size:1.5rem;text-shadow:0 1px 2px rgba(0,0,0,.18)}.rank-inside .total-highlight{background:rgba(15,23,42,.22);border-radius:8px}.rank-inside .total-highlight strong{font-size:1.05rem}
.table-wrap{overflow:auto;max-height:590px;border:1px solid var(--line);border-radius:12px;background:white}.report{border-collapse:separate;border-spacing:0;white-space:nowrap;width:100%;font-size:.72rem}.report th{position:sticky;top:0;background:#0F172A;color:white;padding:8px 9px;z-index:2;font-size:.65rem}.report td{border-right:1px solid #EEF2F7;border-bottom:1px solid #EEF2F7;padding:6px 8px;text-align:center}.report tr:hover td{background-color:#F8FAFC}.empty-bi{background:white;border:1px dashed #CBD5E1;border-radius:12px;padding:22px;text-align:center;color:#64748B}
.stButton button,.stDownloadButton button{border-radius:9px;background:#0F172A;color:white;border:0;font-weight:750;min-height:40px}.stButton button:hover,.stDownloadButton button:hover{background:#1E293B;color:white;border:0}[data-testid="stPopover"] button{border-radius:10px!important;background:#0F172A!important;color:white!important;border:1px solid #334155!important;min-width:48px!important;font-size:1.35rem!important}[data-testid="stDataFrame"]{border:1px solid var(--line);border-radius:12px;overflow:hidden}.element-container{margin-bottom:.2rem}
@media(max-width:1250px){.block-container{padding:.7rem}.rank-row{grid-template-columns:45px 1fr}.rank-name{grid-template-columns:1fr}.rank-inside{grid-template-columns:repeat(4,1fr)}.metric strong{font-size:1.35rem}}@media(max-width:720px){.bi-topbar{padding:14px}.bi-topbar h1{font-size:1.05rem}.rank-row{grid-template-columns:38px 1fr;padding:7px 4px}.rank-inside{grid-template-columns:repeat(2,1fr)}.block-container{padding:.5rem}}

.exec-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:8px 0 14px}.exec-card{background:white;border:1px solid var(--line);border-radius:14px;padding:14px 16px;box-shadow:0 2px 10px rgba(15,23,42,.05);min-width:0}.exec-card small{display:block;font-size:.64rem;font-weight:900;letter-spacing:.055em;color:#64748B}.exec-card strong{display:block;font-size:1.8rem;line-height:1.05;margin-top:7px;color:#0F172A;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.exec-card span{display:block;font-size:.67rem;color:#94A3B8;margin-top:5px}.exec-card.hero{grid-column:span 2;background:linear-gradient(120deg,#0F172A,#172554);border-color:#172554}.exec-card.hero small,.exec-card.hero strong,.exec-card.hero span{color:white}.exec-card.hero strong{font-size:2.35rem}.exec-card.projection{border-top:4px solid #F59E0B}.exec-card.attainment{border-top:4px solid #22C55E}.exec-card.critical{border-top:4px solid #EF4444}.seller-groups{margin-top:12px}.seller-group-title{font-size:.72rem;font-weight:900;letter-spacing:.08em;color:#475569;margin:14px 0 7px}.seller-group-title.award{margin-top:18px}.seller-kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:9px}.seller-kpi{background:white;border:1px solid var(--line);border-radius:12px;padding:12px;min-width:0}.seller-kpi small{display:block;font-size:.59rem;font-weight:900;color:#64748B;letter-spacing:.035em}.seller-kpi strong{display:block;font-size:1.2rem;color:#0F172A;margin-top:7px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.seller-kpi span{display:block;font-size:.63rem;color:#94A3B8;margin-top:4px}.seller-kpi.primary{background:#0F172A;border-color:#0F172A}.seller-kpi.primary small,.seller-kpi.primary strong,.seller-kpi.primary span{color:white}.seller-kpi.primary strong{font-size:1.55rem}.seller-kpi.total{background:linear-gradient(120deg,#0F172A,#172554)}
@media(max-width:900px){.exec-grid{grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}.exec-card.hero{grid-column:span 2}.exec-card.hero strong{font-size:2rem}.seller-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.rank-inside{grid-template-columns:repeat(2,minmax(0,1fr))}.rank-name{grid-template-columns:1fr}.rank-inside strong{white-space:normal;overflow-wrap:anywhere}.report{font-size:.68rem}}
@media(max-width:560px){.block-container{padding:.45rem!important}.bi-topbar{border-radius:10px;padding:12px}.exec-grid{grid-template-columns:repeat(2,minmax(0,1fr));gap:7px}.exec-card{padding:11px 10px;border-radius:11px}.exec-card.hero{grid-column:span 2}.exec-card strong{font-size:1.35rem}.exec-card.hero strong{font-size:1.9rem}.exec-card small{font-size:.56rem}.exec-card span{font-size:.58rem}.seller-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr));gap:7px}.seller-kpi{padding:10px 9px}.seller-kpi strong{font-size:1.05rem}.seller-kpi.primary strong{font-size:1.3rem}.rank-row{grid-template-columns:34px minmax(0,1fr)!important;padding:6px 2px!important}.rank-pos{font-size:.8rem}.rank-name{padding:9px!important;gap:8px!important}.rank-inside{grid-template-columns:repeat(2,minmax(0,1fr))!important}.rank-inside span{min-width:0}.rank-inside strong{font-size:.86rem!important;white-space:normal!important;overflow-wrap:anywhere}.rank-inside .neo-highlight strong{font-size:1.05rem!important}.rank-seller b{white-space:normal!important}.metric strong{white-space:normal;overflow-wrap:anywhere}.stDownloadButton button{width:100%}}

.seller-mobile-primary{display:none}
@media(max-width:560px){.seller-mobile-primary{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:7px;margin-top:12px}.seller-kpi-grid .mobile-duplicate{display:none}.seller-mobile-primary .seller-kpi{min-width:0}.seller-mobile-primary .seller-kpi strong{white-space:normal;overflow-wrap:anywhere}.seller-groups{margin-top:8px}}


.perf-summary{background:white;border:1px solid var(--line);border-radius:14px;padding:12px;display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;box-shadow:0 2px 10px rgba(15,23,42,.04)}.perf-chip{position:relative;border:1px solid #E2E8F0;border-radius:11px;padding:12px 14px;background:#F8FAFC;overflow:hidden}.perf-chip:before{content:"";position:absolute;left:0;top:0;bottom:0;width:5px;background:var(--chip)}.perf-chip span{display:block;font-size:.64rem;font-weight:900;color:#64748B;text-transform:uppercase;letter-spacing:.05em}.perf-chip strong{display:block;font-size:1.65rem;margin-top:6px;color:#0F172A}.channel-summary{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}.channel-group{background:white;border:1px solid var(--line);border-radius:14px;padding:12px;display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px;box-shadow:0 2px 10px rgba(15,23,42,.04)}.channel-mini{border:1px solid #E2E8F0;border-radius:11px;padding:13px 14px;background:#F8FAFC;min-width:0}.channel-mini span{display:block;font-size:.64rem;font-weight:900;color:#64748B;letter-spacing:.04em}.channel-mini strong{display:block;font-size:1.75rem;line-height:1.05;margin-top:7px;color:#0F172A}.channel-mini small{display:block;font-size:.66rem;color:#94A3B8;margin-top:5px}
@media(max-width:720px){.perf-summary{grid-template-columns:repeat(2,minmax(0,1fr));gap:7px;padding:8px}.perf-chip{padding:10px 11px}.perf-chip strong{font-size:1.4rem}.channel-summary{grid-template-columns:1fr;gap:8px}.channel-group{grid-template-columns:repeat(2,minmax(0,1fr));gap:7px;padding:8px}.channel-mini{padding:10px}.channel-mini strong{font-size:1.4rem}.rank-card{border-radius:12px}.rank-row{grid-template-columns:28px minmax(0,1fr)!important;gap:4px!important;padding:5px 3px!important}.rank-pos{font-size:.72rem}.rank-name{padding:9px!important;border-radius:10px!important;gap:7px!important}.rank-seller{padding-top:0!important;margin-bottom:3px}.rank-seller b{font-size:.9rem!important;line-height:1.15}.rank-seller small{font-size:.57rem!important}.rank-inside{display:grid!important;grid-template-columns:repeat(6,minmax(0,1fr))!important;gap:5px!important}.rank-inside span{grid-column:span 2;min-height:42px!important;padding:5px 3px!important;border-left:0!important;border:1px solid rgba(255,255,255,.18);border-radius:7px;background:rgba(255,255,255,.06)}.rank-inside .main-kpi{grid-column:span 3!important;background:rgba(15,23,42,.20)}.rank-inside .main-kpi strong{font-size:1.55rem!important}.rank-inside .main-kpi small{font-size:.58rem!important;font-weight:800}.rank-inside strong{font-size:.82rem!important;line-height:1.05!important}.rank-inside small{font-size:.49rem!important;margin-top:3px!important}.rank-inside .neo-highlight{grid-column:span 2!important}.rank-inside .total-highlight{grid-column:span 3!important}.rank-inside .total-highlight strong{font-size:.98rem!important}}


.projection-status{display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;background:white;border:1px solid var(--line);box-shadow:none;min-height:100%}.projection-status-emoji{font-size:2.25rem;line-height:1;margin:1px 0 8px}.projection-status-message{font-size:.70rem;line-height:1.15;font-weight:900;letter-spacing:.035em;color:#475569;text-align:center;word-break:normal;overflow-wrap:break-word}
@media(max-width:560px){.projection-status{padding:9px 7px!important;min-height:82px}.projection-status-emoji{font-size:1.8rem;margin-bottom:6px}.projection-status-message{font-size:.58rem;line-height:1.12}}


.rank-inside .rank-projection-status{background:rgba(255,255,255,.12);border-radius:8px;border-left:0}.rank-inside .rank-projection-status .rank-status-emoji{font-size:1.7rem;line-height:1}.rank-inside .rank-projection-status small{font-size:.52rem;font-weight:900;line-height:1.05;text-align:center}
@media(max-width:720px){.rank-inside .rank-projection-status{grid-column:span 2!important}.rank-inside .rank-projection-status .rank-status-emoji{font-size:1.45rem!important}.rank-inside .rank-projection-status small{font-size:.46rem!important;line-height:1.05!important}}


.bi-topbar-nav{display:flex;align-items:center;justify-content:space-between;gap:24px}.bi-brand{min-width:0;flex:1}.top-nav{display:flex;align-items:center;justify-content:flex-end;gap:6px;flex-wrap:wrap}.top-nav-item{display:inline-flex;align-items:center;justify-content:center;padding:9px 12px;border-radius:9px;color:#CBD5E1!important;text-decoration:none!important;font-size:.68rem;font-weight:800;letter-spacing:.02em;white-space:nowrap;border:1px solid transparent;transition:.15s ease}.top-nav-item:hover{background:rgba(255,255,255,.09);color:white!important}.top-nav-item.active{background:white;color:#0F172A!important;border-color:rgba(255,255,255,.7);box-shadow:0 2px 8px rgba(0,0,0,.12)}
@media(max-width:980px){.bi-topbar-nav{align-items:flex-start;flex-direction:column;gap:12px}.top-nav{justify-content:flex-start;width:100%}.top-nav-item{padding:8px 10px;font-size:.62rem}}
@media(max-width:560px){.bi-topbar-nav{gap:10px}.top-nav{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:6px}.top-nav-item{width:100%;padding:8px 6px;font-size:.58rem}.top-nav-item:last-child{grid-column:span 2}.bi-brand p{font-size:.68rem!important}}


/* Navegação nativa: troca apenas a área da aplicação, sem abrir nova página/aba. */
[data-testid="stSegmentedControl"]{margin-top:-18px;margin-bottom:10px;background:#172554;border-radius:0 0 14px 14px;padding:0 16px 14px}
[data-testid="stSegmentedControl"] button{color:#CBD5E1!important;border-color:rgba(255,255,255,.16)!important;font-weight:800!important;font-size:.68rem!important}
[data-testid="stSegmentedControl"] button[aria-pressed="true"]{background:#FFFFFF!important;color:#0F172A!important;border-color:#FFFFFF!important}
[data-testid="stSegmentedControl"] button:hover{background:rgba(255,255,255,.10)!important;color:white!important}
@media(max-width:720px){[data-testid="stSegmentedControl"]{padding:0 8px 10px;overflow-x:auto}[data-testid="stSegmentedControl"]>div{min-width:max-content}[data-testid="stSegmentedControl"] button{font-size:.58rem!important;padding-left:8px!important;padding-right:8px!important}}


.rank-click{display:block;color:inherit!important;text-decoration:none!important;min-width:0}.rank-click:hover .rank-name{filter:brightness(1.03);box-shadow:0 4px 12px rgba(15,23,42,.18);transform:translateY(-1px)}.rank-name{transition:filter .12s ease,box-shadow .12s ease,transform .12s ease}.login-shell{min-height:58vh;display:flex;align-items:center;justify-content:center;padding:32px 12px 10px}.login-card{width:min(520px,100%);background:linear-gradient(120deg,#0F172A,#172554);border-radius:18px;padding:28px 26px;color:white;box-shadow:0 14px 44px rgba(15,23,42,.18);margin-bottom:4px}.login-brand{font-weight:900;font-size:1.35rem;letter-spacing:-.02em}.login-title{font-size:1rem;font-weight:800;margin-top:24px}.login-sub{color:#CBD5E1;font-size:.78rem;margin-top:5px}
@media(max-width:560px){.login-shell{min-height:36vh;padding:18px 4px 4px}.login-card{padding:22px 18px;border-radius:14px}.login-brand{font-size:1.05rem}.rank-click{width:100%}[data-testid="stDialog"] [role="dialog"]{max-width:calc(100vw - 14px)!important;width:calc(100vw - 14px)!important;max-height:92vh!important}[data-testid="stDialog"] .seller-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important}[data-testid="stDialog"] .seller-mobile-primary{display:grid!important}}


.seller-dialog-meta{display:flex;align-items:center;justify-content:space-between;gap:8px;margin:-4px 0 8px;padding:7px 10px;border-radius:9px;background:#F8FAFC;border-left:4px solid var(--seller-color);font-size:.72rem;color:#64748B}.seller-dialog-meta b{color:#334155;font-size:.7rem}
@media(max-width:560px){
[data-testid="stDialog"] [role="dialog"]{width:calc(100vw - 12px)!important;max-width:calc(100vw - 12px)!important;max-height:94dvh!important;margin:3dvh auto!important;overflow:hidden!important}
[data-testid="stDialog"] [role="dialog"]>div{max-height:94dvh!important;overflow-y:auto!important;padding:.65rem .65rem .8rem!important}
[data-testid="stDialog"] h2{font-size:1.1rem!important;line-height:1.1!important;margin:0 0 .3rem!important}
[data-testid="stDialog"] .seller-dialog-meta{margin:-2px 0 6px;padding:5px 8px;font-size:.62rem}.seller-dialog-meta b{font-size:.62rem}
[data-testid="stDialog"] .seller-mobile-primary{display:grid!important;grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:6px!important;margin:0 0 7px!important}
[data-testid="stDialog"] .seller-mobile-primary .seller-kpi{min-height:70px!important;padding:8px!important;border-radius:10px!important}
[data-testid="stDialog"] .seller-mobile-primary .seller-kpi strong{font-size:1.18rem!important;margin-top:4px!important}
[data-testid="stDialog"] .seller-mobile-primary .seller-kpi small{font-size:.53rem!important;line-height:1.05!important}
[data-testid="stDialog"] .seller-mobile-primary .seller-kpi span{font-size:.54rem!important;margin-top:3px!important}
[data-testid="stDialog"] .seller-groups{margin-top:4px!important}
[data-testid="stDialog"] .seller-group-title{font-size:.61rem!important;margin:7px 0 4px!important;letter-spacing:.05em!important}
[data-testid="stDialog"] .seller-group-title.award{margin-top:8px!important}
[data-testid="stDialog"] .seller-kpi-grid{grid-template-columns:repeat(3,minmax(0,1fr))!important;gap:5px!important}
[data-testid="stDialog"] .seller-kpi-grid .seller-kpi{min-height:52px!important;padding:6px 5px!important;border-radius:9px!important}
[data-testid="stDialog"] .seller-kpi-grid .seller-kpi strong{font-size:.86rem!important;margin-top:3px!important;white-space:normal!important;line-height:1.05!important}
[data-testid="stDialog"] .seller-kpi-grid .seller-kpi small{font-size:.47rem!important;line-height:1.05!important;letter-spacing:.015em!important}
[data-testid="stDialog"] .seller-kpi-grid .seller-kpi span{display:none!important}
[data-testid="stDialog"] .mobile-duplicate{display:none!important}
[data-testid="stDialog"] .stButton{display:none!important}
}


.integrated-header{display:flex;align-items:center;justify-content:space-between;gap:18px}.header-account{display:flex;align-items:center;gap:14px;flex:0 0 auto}.header-user{font-size:.74rem;font-weight:800;color:#E2E8F0;white-space:nowrap}.header-actions{display:flex;gap:7px}.header-actions a{display:inline-flex;align-items:center;justify-content:center;min-height:34px;padding:0 11px;border:1px solid rgba(255,255,255,.24);border-radius:8px;color:white!important;text-decoration:none!important;font-size:.65rem;font-weight:900;letter-spacing:.03em}.header-actions a:hover{background:rgba(255,255,255,.10)}
.rank-mobile-status{display:none}.rank-mobile-head{display:contents}
@media(max-width:560px){
.integrated-header{padding:11px 12px!important;gap:8px!important;align-items:flex-start!important}.integrated-header .bi-brand{min-width:0;flex:1}.integrated-header .bi-brand h1{font-size:.95rem!important;white-space:normal}.integrated-header .bi-brand p{display:none}.header-account{flex-direction:column;align-items:flex-end;gap:5px;max-width:44%}.header-user{font-size:.58rem;max-width:100%;overflow:hidden;text-overflow:ellipsis}.header-actions{gap:4px}.header-actions a{min-height:27px;padding:0 7px;font-size:.52rem;border-radius:6px}
.rank-card{overflow:visible!important}.rank-row{grid-template-columns:30px minmax(0,1fr)!important;gap:5px!important;padding:5px 2px!important}.rank-pos{font-size:.76rem}.rank-click{min-width:0!important;width:100%!important}.rank-name{display:block!important;padding:9px!important;border-radius:10px!important;min-width:0!important;width:100%!important;box-sizing:border-box!important}.rank-seller{padding:0!important}.rank-mobile-head{display:flex!important;align-items:flex-start!important;justify-content:space-between!important;gap:8px!important;margin-bottom:7px}.rank-mobile-head>div:first-child{min-width:0;flex:1}.rank-seller b{font-size:.79rem!important;line-height:1.05!important;white-space:normal!important;overflow:visible!important;text-overflow:clip!important}.rank-seller small{font-size:.53rem!important;margin-top:3px!important}.rank-mobile-status{display:flex!important;flex-direction:column!important;align-items:center!important;justify-content:center!important;min-width:60px!important;text-align:center!important}.rank-mobile-status strong{font-size:1.35rem!important;line-height:1!important}.rank-mobile-status small{font-size:.43rem!important;line-height:1.05!important;margin-top:3px!important;color:rgba(255,255,255,.92)!important;font-weight:800!important}.rank-inside{display:grid!important;grid-template-columns:repeat(3,minmax(0,1fr))!important;gap:4px!important}.rank-inside span{min-width:0!important;min-height:42px!important;padding:4px 3px!important;border-left:0!important;background:rgba(255,255,255,.07);border-radius:7px}.rank-inside strong{font-size:.72rem!important;white-space:normal!important;overflow-wrap:anywhere!important}.rank-inside small{font-size:.42rem!important;line-height:1.05!important;margin-top:3px!important}.rank-inside .main-kpi{grid-column:span 1;background:rgba(15,23,42,.18)!important;min-height:52px!important}.rank-inside .main-kpi strong{font-size:1.16rem!important}.rank-inside .neo-highlight strong{font-size:.9rem!important}.rank-inside .total-highlight{grid-column:span 3!important;min-height:48px!important}.rank-inside .total-highlight strong{font-size:1.02rem!important}.rank-inside .desktop-status{display:none!important}
[data-testid="stDialog"] [role="dialog"]{box-sizing:border-box!important;width:calc(100vw - 16px)!important;max-width:calc(100vw - 16px)!important;margin-left:auto!important;margin-right:auto!important;overflow-x:clip!important}[data-testid="stDialog"] [role="dialog"]>div{box-sizing:border-box!important;width:100%!important;max-width:100%!important;padding:.58rem!important;overflow-x:clip!important}[data-testid="stDialog"] .seller-mobile-primary,[data-testid="stDialog"] .seller-kpi-grid{width:100%!important;max-width:100%!important;box-sizing:border-box!important}[data-testid="stDialog"] .seller-mobile-primary{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:5px!important}[data-testid="stDialog"] .seller-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:5px!important}[data-testid="stDialog"] .seller-kpi{box-sizing:border-box!important;min-width:0!important;width:100%!important;max-width:100%!important;padding:7px!important}[data-testid="stDialog"] .seller-kpi strong{white-space:normal!important;overflow-wrap:anywhere!important;word-break:normal!important}[data-testid="stDialog"] .seller-dialog-meta{box-sizing:border-box!important;width:100%!important;max-width:100%!important}
}
@media(max-width:350px){[data-testid="stDialog"] .seller-kpi-grid{grid-template-columns:1fr 1fr!important}.rank-inside{grid-template-columns:repeat(2,minmax(0,1fr))!important}.rank-inside .total-highlight{grid-column:span 2!important}}


.rank-mobile-head>div:first-child{display:flex;flex-direction:column;min-width:0}


/* Ajustes finais exclusivamente mobile: ordem lógica do ranking e dialog sem overflow. */
@media(max-width:560px){
/* Ranking: 6 colunas virtuais para preservar a sequência Vendas/Projeção -> Ritmo -> Neo -> Premiação -> Total. */
.rank-inside{grid-template-columns:repeat(6,minmax(0,1fr))!important;gap:4px!important}
.rank-inside>span:nth-child(1),.rank-inside>span:nth-child(2){grid-column:span 3!important;min-height:51px!important;background:rgba(15,23,42,.18)!important}
.rank-inside>span:nth-child(1) strong,.rank-inside>span:nth-child(2) strong{font-size:1.16rem!important}
.rank-inside>span:nth-child(3),.rank-inside>span:nth-child(4),.rank-inside>span:nth-child(5){grid-column:span 2!important}
.rank-inside>span:nth-child(6),.rank-inside>span:nth-child(7){grid-column:span 3!important;background:rgba(255,255,255,.13)!important}
.rank-inside>span:nth-child(8),.rank-inside>span:nth-child(9),.rank-inside>span:nth-child(10),.rank-inside>span:nth-child(11){grid-column:span 3!important}
.rank-inside>span:nth-child(12){grid-column:span 6!important}
.rank-inside>span:nth-child(13){grid-column:span 6!important;min-height:48px!important;background:rgba(15,23,42,.24)!important}
.rank-inside>span:nth-child(14){display:none!important}
.rank-inside>span{box-sizing:border-box!important;min-width:0!important;max-width:100%!important}
.rank-inside strong{max-width:100%!important;overflow-wrap:anywhere!important}

/* Dialog do vendedor: o conteúdo é dimensionado para caber de verdade na viewport. */
[data-testid="stDialog"] [role="dialog"]{box-sizing:border-box!important;width:calc(100vw - 16px)!important;max-width:calc(100vw - 16px)!important;min-width:0!important;max-height:94dvh!important;margin:3dvh auto!important}
[data-testid="stDialog"] [role="dialog"]>div{box-sizing:border-box!important;width:100%!important;max-width:100%!important;min-width:0!important;max-height:94dvh!important;overflow-y:auto!important;padding:.58rem!important}
[data-testid="stDialog"] [role="dialog"] *,[data-testid="stDialog"] .seller-mobile-primary,[data-testid="stDialog"] .seller-kpi-grid,[data-testid="stDialog"] .seller-kpi,[data-testid="stDialog"] .seller-dialog-meta{box-sizing:border-box!important;min-width:0!important;max-width:100%!important}
[data-testid="stDialog"] .seller-mobile-primary,[data-testid="stDialog"] .seller-kpi-grid{width:100%!important;grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:5px!important}
[data-testid="stDialog"] .seller-mobile-primary{margin:0 0 6px!important}
[data-testid="stDialog"] .seller-kpi{width:100%!important;padding:7px 6px!important;border-radius:9px!important}
[data-testid="stDialog"] .seller-kpi strong{font-size:clamp(.76rem,3.25vw,.91rem)!important;line-height:1.06!important;white-space:normal!important;overflow-wrap:anywhere!important;word-break:normal!important}
[data-testid="stDialog"] .seller-mobile-primary .seller-kpi strong{font-size:clamp(.98rem,4.4vw,1.18rem)!important}
[data-testid="stDialog"] .seller-kpi small{font-size:.46rem!important;line-height:1.05!important}
[data-testid="stDialog"] .seller-dialog-meta{width:100%!important;margin:-2px 0 5px!important;padding:5px 7px!important}
[data-testid="stDialog"] .seller-group-title{margin:6px 0 4px!important}
}
@media(max-width:340px){
[data-testid="stDialog"] [role="dialog"]{width:calc(100vw - 12px)!important;max-width:calc(100vw - 12px)!important}
[data-testid="stDialog"] .seller-kpi-grid,[data-testid="stDialog"] .seller-mobile-primary{gap:4px!important}
[data-testid="stDialog"] .seller-kpi{padding:6px 5px!important}
.rank-inside strong{font-size:.68rem!important}.rank-inside small{font-size:.38rem!important}
}


/* AJUSTE FINAL — EXCLUSIVAMENTE RANKING MOBILE. Desktop permanece intocado. */
@media(max-width:560px){
  /* reduz altura pelo layout, não pela perda de legibilidade */
  .rank-card{border-radius:11px!important}
  .rank-row{grid-template-columns:25px minmax(0,1fr)!important;gap:3px!important;padding:3px 1px!important}
  .rank-pos{font-size:.76rem!important;line-height:1!important}
  .rank-name{display:block!important;width:100%!important;min-width:0!important;box-sizing:border-box!important;padding:6px 7px!important;border-radius:9px!important;gap:0!important}
  .rank-seller{padding:0!important}
  .rank-mobile-head{display:flex!important;align-items:center!important;justify-content:space-between!important;gap:5px!important;margin:0 0 4px!important;min-height:24px!important}
  .rank-mobile-head>div:first-child{min-width:0!important;flex:1!important}
  .rank-seller b{font-size:.82rem!important;line-height:1.02!important;font-weight:900!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}
  .rank-seller small{font-size:.52rem!important;line-height:1!important;margin-top:2px!important}
  .rank-mobile-status{display:flex!important;flex-direction:row!important;align-items:center!important;justify-content:flex-end!important;gap:3px!important;min-width:0!important;max-width:34%!important;text-align:right!important}
  .rank-mobile-status strong{font-size:1.02rem!important;line-height:1!important}
  .rank-mobile-status small{font-size:.42rem!important;line-height:1!important;margin:0!important;font-weight:850!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;color:rgba(255,255,255,.94)!important}

  /* 12 colunas permitem concentrar todas as informações em 3 linhas compactas. */
  .rank-inside{display:grid!important;grid-template-columns:repeat(12,minmax(0,1fr))!important;gap:2px!important;align-items:stretch!important}
  .rank-inside span{min-width:0!important;min-height:25px!important;padding:2px 2px!important;border-left:0!important;border-radius:5px!important;background:rgba(255,255,255,.065)!important;box-sizing:border-box!important;text-align:center!important}
  .rank-inside strong{font-size:.66rem!important;line-height:1!important;font-weight:850!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;max-width:100%!important}
  .rank-inside small{font-size:.34rem!important;line-height:1!important;margin-top:2px!important;letter-spacing:0!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}

  /* LINHA 1 — prioridade máxima: Vendas, Projeção, Premiação atual e projetada. */
  .rank-inside span:nth-child(1),
  .rank-inside span:nth-child(2),
  .rank-inside span:nth-child(8),
  .rank-inside span:nth-child(9){grid-column:span 3!important;grid-row:1!important;min-height:34px!important;background:rgba(15,23,42,.20)!important}
  .rank-inside span:nth-child(1) strong,
  .rank-inside span:nth-child(2) strong{font-size:1.04rem!important;font-weight:950!important}
  .rank-inside span:nth-child(8) strong,
  .rank-inside span:nth-child(9) strong{font-size:.78rem!important;font-weight:950!important}
  .rank-inside span:nth-child(1) small,
  .rank-inside span:nth-child(2) small,
  .rank-inside span:nth-child(8) small,
  .rank-inside span:nth-child(9) small{font-size:.36rem!important;font-weight:850!important}

  /* LINHA 2 — ritmo, meta, Neo e total variável. */
  .rank-inside span:nth-child(13){grid-column:span 4!important;grid-row:2!important;min-height:29px!important;background:rgba(15,23,42,.25)!important}
  .rank-inside span:nth-child(13) strong{font-size:.78rem!important;font-weight:950!important}
  .rank-inside span:nth-child(3),
  .rank-inside span:nth-child(5),
  .rank-inside span:nth-child(6),
  .rank-inside span:nth-child(7){grid-column:span 2!important;grid-row:2!important;min-height:29px!important}
  .rank-inside span:nth-child(6),.rank-inside span:nth-child(7){background:rgba(255,255,255,.13)!important}
  .rank-inside span:nth-child(6) strong,.rank-inside span:nth-child(7) strong{font-size:.72rem!important}

  /* LINHA 3 — complementares financeiros e zeros, sem esconder nenhum dado. */
  .rank-inside span:nth-child(4){grid-column:span 2!important;grid-row:3!important;min-height:26px!important}
  .rank-inside span:nth-child(10){grid-column:span 3!important;grid-row:3!important;min-height:26px!important}
  .rank-inside span:nth-child(11){grid-column:span 4!important;grid-row:3!important;min-height:26px!important}
  .rank-inside span:nth-child(12){grid-column:span 3!important;grid-row:3!important;min-height:26px!important}

  /* Status separado do desktop não ocupa espaço no mobile; emoji/status já estão no cabeçalho. */
  .rank-inside .desktop-status{display:none!important}
}

@media(max-width:350px){
  .rank-row{grid-template-columns:23px minmax(0,1fr)!important}
  .rank-name{padding:5px 6px!important}
  .rank-seller b{font-size:.76rem!important}
  .rank-inside strong{font-size:.61rem!important}
  .rank-inside span:nth-child(1) strong,.rank-inside span:nth-child(2) strong{font-size:.96rem!important}
  .rank-inside span:nth-child(8) strong,.rank-inside span:nth-child(9) strong,.rank-inside span:nth-child(13) strong{font-size:.70rem!important}
}


.pc-dashboard{background:#fff;border:1px solid var(--line);border-radius:14px;padding:14px 16px;box-shadow:0 2px 10px rgba(15,23,42,.04);font-family:inherit}.pc-channel-row{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px;margin-bottom:12px}.pc-channel-card{background:#F8FAFC;border:1px solid var(--line);border-radius:11px;padding:12px 14px;min-width:0}.pc-channel-content{min-width:0}.pc-channel-content span,.pc-team-title,.pc-team-stats span,.pc-team-sales span,.pc-team-progress span,.pc-team-goal span{font-family:inherit;font-size:.68rem;font-weight:850;letter-spacing:.035em;color:#64748B}.pc-channel-content strong{display:block;font-family:inherit;font-size:1.75rem;line-height:1;color:#0F172A;margin:7px 0 4px;font-weight:950}.pc-channel-content small{font-family:inherit;font-size:.66rem;color:#94A3B8}.pc-channel-track{height:5px;background:#E5E7EB;border-radius:99px;margin-top:8px;overflow:hidden}.pc-channel-track i{display:block;height:100%;border-radius:99px;background:#4F6FE8}.pc-channel-card.green .pc-channel-track i{background:#2E9D55}.pc-team-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}.pc-team-card{border:1px solid var(--line);border-top:4px solid #6D4BC3;border-radius:12px;padding:13px 14px;background:#fff;min-width:0;box-shadow:0 2px 10px rgba(15,23,42,.035)}.pc-team-card.external{border-top-color:#24974B}.pc-team-head{display:flex;align-items:center;justify-content:space-between;gap:12px}.pc-team-title{display:flex;align-items:center;color:#475569}.pc-team-title b{font-size:.74rem;font-weight:900}.pc-team-goal{display:flex;align-items:center;gap:8px;background:#F1F5F9;border:1px solid #CBD5E1;border-radius:10px;padding:7px 10px;white-space:nowrap}.pc-team-goal span{font-size:.64rem}.pc-team-goal b{font-family:inherit;font-size:1.45rem;line-height:1;color:#0F172A;font-weight:950}.pc-team-main{display:grid;grid-template-columns:1fr .75fr;align-items:center;gap:16px;padding:14px 2px 13px;border-bottom:1px solid #E8EDF4}.pc-team-sales{display:flex;align-items:flex-end;gap:8px}.pc-team-sales strong{font-family:inherit;font-size:2.3rem;line-height:.95;color:#0F172A;font-weight:950}.pc-team-sales span{padding-bottom:4px}.pc-team-progress{text-align:right}.pc-team-progress b{font-family:inherit;font-size:1.05rem;color:#0F172A;font-weight:900}.pc-team-progress span{margin-left:4px;font-size:.62rem}.pc-team-track{height:6px;background:#E8EBF0;border-radius:99px;margin-top:8px;overflow:hidden}.pc-team-track i{display:block;height:100%;background:#6D4BC3;border-radius:99px}.pc-team-card.external .pc-team-track i{background:#24974B}.pc-team-stats{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));padding-top:12px}.pc-team-stats>div{text-align:center;padding:0 7px;border-right:1px solid #E2E8F0;min-width:0}.pc-team-stats>div:last-child{border-right:0}.pc-team-stats span{display:block;white-space:normal;font-size:.60rem}.pc-team-stats strong{display:block;font-family:inherit;font-size:1.15rem;color:#0F172A;margin:6px 0 0;font-weight:900}@media(max-width:700px){.pc-dashboard{padding:10px}.pc-channel-row{grid-template-columns:1fr 1fr;gap:7px;margin-bottom:9px}.pc-channel-card{padding:10px}.pc-channel-content span,.pc-team-title,.pc-team-stats span,.pc-team-sales span,.pc-team-progress span,.pc-team-goal span{font-size:.58rem}.pc-channel-content strong{font-size:1.42rem}.pc-team-grid{grid-template-columns:1fr;gap:8px}.pc-team-card{padding:10px}.pc-team-title b{font-size:.68rem}.pc-team-goal{padding:6px 8px}.pc-team-goal b{font-size:1.28rem}.pc-team-sales strong{font-size:2rem}.pc-team-progress b{font-size:.96rem}.pc-team-stats strong{font-size:1.03rem}}@media(max-width:360px){.pc-team-stats{grid-template-columns:repeat(2,1fr);row-gap:10px}.pc-team-stats>div:nth-child(2){border-right:0}.pc-team-goal b{font-size:1.18rem}}
.weekly-rank-list{display:grid;gap:7px;margin:8px 0 12px}.weekly-rank-card{display:grid;grid-template-columns:42px minmax(160px,1fr) 120px 160px;align-items:center;gap:10px;background:white;border:1px solid var(--line);border-radius:11px;padding:10px 13px;box-shadow:0 2px 8px rgba(15,23,42,.035)}.weekly-rank-pos{font-weight:900;color:#475569;text-align:center}.weekly-rank-name{font-size:.78rem;font-weight:900;color:#0F172A;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.weekly-rank-sales,.weekly-rank-award{display:flex;flex-direction:column;align-items:flex-end}.weekly-rank-sales strong{font-size:1.25rem;color:#0F172A}.weekly-rank-award strong{font-size:1rem;color:#0F172A}.weekly-rank-sales span,.weekly-rank-award span{font-size:.52rem;font-weight:850;color:#64748B;margin-top:2px}.weekly-history{background:white;border:1px solid var(--line);border-radius:12px;padding:12px 14px;margin:8px 0 14px}.weekly-history-name{font-size:.82rem;font-weight:900;color:#0F172A;margin-bottom:8px}.weekly-history-row,.weekly-history-total{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;align-items:center;padding:8px 0;border-top:1px solid #EEF2F7}.weekly-history-row span,.weekly-history-total span{font-size:.6rem;font-weight:850;color:#64748B}.weekly-history-row strong,.weekly-history-total strong{font-size:.76rem;color:#0F172A}.weekly-history-row b,.weekly-history-total b{text-align:right;font-size:.76rem;color:#0F172A}.weekly-history-total{background:#F8FAFC;margin:5px -6px -4px;padding:10px 6px;border-radius:8px}.weekly-desktop-table{overflow:auto;border:1px solid var(--line);border-radius:12px;background:white;margin-top:12px}.weekly-desktop-table table{border-collapse:collapse;width:100%;white-space:nowrap;font-size:.68rem}.weekly-desktop-table th{background:#0F172A;color:white;padding:8px;font-size:.6rem}.weekly-desktop-table td{padding:8px;border-right:1px solid #EEF2F7;border-bottom:1px solid #EEF2F7;text-align:center}.weekly-desktop-table td strong{display:block;font-size:.72rem}.weekly-desktop-table td small{font-size:.5rem;color:#94A3B8}.weekly-desktop-table .weekly-name{text-align:left;font-weight:850;position:sticky;left:0;background:white;z-index:1}
@media(max-width:700px){.weekly-rank-list{gap:5px}.weekly-rank-card{grid-template-columns:28px minmax(0,1fr) 74px 104px;gap:5px;padding:7px 8px;border-radius:9px}.weekly-rank-pos{font-size:.68rem}.weekly-rank-name{font-size:.72rem}.weekly-rank-sales strong{font-size:1.06rem}.weekly-rank-award strong{font-size:.76rem}.weekly-rank-sales span,.weekly-rank-award span{font-size:.43rem}.weekly-history{padding:10px}.weekly-history-row,.weekly-history-total{grid-template-columns:.8fr 1fr 1fr;gap:6px;padding:7px 0}.weekly-history-row span,.weekly-history-total span{font-size:.52rem}.weekly-history-row strong,.weekly-history-total strong,.weekly-history-row b,.weekly-history-total b{font-size:.68rem}.weekly-desktop-table{display:none}}

.weekly-game-list{display:grid;gap:7px;margin:9px 0 12px}.weekly-game-card{background:#fff;border:1px solid var(--line);border-radius:12px;padding:10px 12px;box-shadow:0 2px 8px rgba(15,23,42,.035)}.weekly-game-head{display:grid;grid-template-columns:34px minmax(0,1fr) auto;align-items:center;gap:7px}.weekly-game-head b{font-size:.78rem;color:#334155;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.weekly-pos{font-size:.72rem;font-weight:950;color:#64748B}.weekly-money-emoji{font-size:1rem;white-space:nowrap;letter-spacing:-1px}.weekly-game-main{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:6px}.weekly-game-main>div{display:flex;align-items:baseline;gap:6px;background:#F8FAFC;border-radius:8px;padding:6px 8px}.weekly-game-main strong{font-size:1.28rem;line-height:1;color:#0F172A;font-weight:950}.weekly-game-main small{font-size:.52rem;color:#64748B;font-weight:900}.weekly-prize strong{font-size:1.04rem}.weekly-target{margin-top:6px;padding:5px 8px;border-radius:7px;background:#FFF7ED;color:#9A3412;font-size:.64rem;font-weight:750}.weekly-target.max{background:#ECFDF5;color:#166534}.weekly-management-table{overflow:auto;max-height:470px;border:1px solid var(--line);border-radius:10px;background:#fff}.weekly-management-table table{border-collapse:collapse;width:100%;white-space:nowrap;font-size:.68rem}.weekly-management-table th{position:sticky;top:0;background:#0F172A;color:#fff;padding:7px 8px;font-size:.60rem;z-index:1}.weekly-management-table td{padding:6px 8px;border-bottom:1px solid #EEF2F7;border-right:1px solid #EEF2F7;text-align:center}.weekly-management-table td:first-child{text-align:left;font-weight:800;color:#334155}@media(max-width:560px){.weekly-game-list{gap:5px}.weekly-game-card{padding:7px 8px;border-radius:10px}.weekly-game-head{grid-template-columns:27px minmax(0,1fr) auto;gap:5px}.weekly-game-head b{font-size:.72rem}.weekly-pos{font-size:.64rem}.weekly-money-emoji{font-size:.84rem}.weekly-game-main{gap:5px;margin-top:4px}.weekly-game-main>div{padding:5px 6px}.weekly-game-main strong{font-size:1.12rem}.weekly-prize strong{font-size:.94rem}.weekly-game-main small{font-size:.46rem}.weekly-target{font-size:.57rem;margin-top:4px;padding:4px 6px}}

</style>"""


def weekly_current_index(cfg,max_weeks,today=None):
    """Índice visual da semana atual sem alterar a regra de cálculo das semanas."""
    today=today or date.today()
    ranges=month_weeks(int(cfg["ano"]),int(cfg["mes"]))[:max_weeks]
    if not ranges:return 0
    if today.year==int(cfg["ano"]) and today.month==int(cfg["mes"]):
        return next((i for i,(a,b) in enumerate(ranges) if a<=today<=b),max_weeks-1)
    # Competência histórica/futura: usa a data de referência apenas para escolher a aba,
    # sem interferir em nenhum cálculo de vendas ou premiação.
    ref_day=max(1,min(int(cfg.get("dia_referencia",1) or 1),ranges[-1][1].day))
    ref=date(int(cfg["ano"]),int(cfg["mes"]),ref_day)
    return next((i for i,(a,b) in enumerate(ranges) if a<=ref<=b),max_weeks-1)

def weekly_tiers(cfg):
    return sorted(cfg.get("premiacao_semanal",[]),key=lambda x:int(x.get("vendas",0)))

def weekly_game_level(sales,cfg):
    tiers=weekly_tiers(cfg)
    conquered=[i for i,t in enumerate(tiers) if sales>=int(t.get("vendas",0)) and float(t.get("premio",0) or 0)>0]
    if not conquered:return 0
    idx=conquered[-1]
    if len(tiers)<=1:return 1
    # Escala proporcional 1..6; primeira faixa=1 e maior faixa=6.
    return max(1,min(6,1+int(idx*5/(len(tiers)-1))))

def weekly_next_goal(sales,cfg):
    tiers=weekly_tiers(cfg)
    nxt=next((t for t in tiers if int(t.get("vendas",0))>sales),None)
    if not nxt:return None
    target=int(nxt.get("vendas",0)); prize=float(nxt.get("premio",0) or 0)
    return {"target":target,"missing":max(0,target-sales),"prize":prize}

def weekly_week_labels(cfg,max_weeks,current_index):
    labels=[]
    for i in range(max_weeks):
        if i<current_index: labels.append(f"✓ SEMANA {i+1}")
        elif i==current_index: labels.append(f"● SEMANA {i+1} — ATUAL")
        else: labels.append(f"SEMANA {i+1}")
    return labels

def weekly_rank_gamified_html(team,week_index,cfg,is_current):
    ranked=sorted(team,key=lambda x:((x.get("semanas",[])[week_index] if week_index<len(x.get("semanas",[])) else 0),(x.get("premios",[])[week_index] if week_index<len(x.get("premios",[])) else 0)),reverse=True)
    rows=[]
    for pos,x in enumerate(ranked,1):
        sales=int(x.get("semanas",[])[week_index] if week_index<len(x.get("semanas",[])) else 0)
        prize=float(x.get("premios",[])[week_index] if week_index<len(x.get("premios",[])) else 0)
        level=weekly_game_level(sales,cfg) if prize>0 else 0
        emojis='🤑'*level
        target=''
        if is_current:
            nxt=weekly_next_goal(sales,cfg)
            if nxt:
                target=f'<div class="weekly-target">🎯 Faltam <b>{nxt["missing"]}</b> vendas para <b>{money(nxt["prize"])}</b></div>'
            else:
                target='<div class="weekly-target max">🏆 Maior premiação atingida</div>'
        rows.append(
            '<div class="weekly-game-card">'
            f'<div class="weekly-game-head"><span class="weekly-pos">{pos}º</span><b>{html.escape(str(x.get("vendedor","")))}</b><span class="weekly-money-emoji">{emojis}</span></div>'
            f'<div class="weekly-game-main"><div><strong>{sales}</strong><small>VENDAS</small></div><div class="weekly-prize"><strong>{money(prize)}</strong><small>SEMANAL</small></div></div>'
            f'{target}</div>'
        )
    return '<div class="weekly-game-list">'+''.join(rows)+'</div>'

def weekly_management_table_html(team,max_weeks):
    heads=['VENDEDOR']
    for i in range(max_weeks):heads += [f'S{i+1} VENDAS',f'S{i+1} VALOR']
    rows=[]
    for x in sorted(team,key=lambda z:normalize_text(z.get("vendedor",""))):
        cells=[html.escape(str(x.get("vendedor","")))]
        for i in range(max_weeks):
            qty=x.get("semanas",[])[i] if i<len(x.get("semanas",[])) else 0
            prize=x.get("premios",[])[i] if i<len(x.get("premios",[])) else 0
            cells += [str(qty),money(prize)]
        rows.append('<tr>'+''.join(f'<td>{c}</td>' for c in cells)+'</tr>')
    return '<div class="weekly-management-table"><table><thead><tr>'+''.join(f'<th>{h}</th>' for h in heads)+'</tr></thead><tbody>'+''.join(rows)+'</tbody></table></div>'


def render_login(st,cfg):
    st.markdown('<div class="login-shell"><div class="login-card"><div class="login-brand">PAINEL COMERCIAL — AFOGADOS</div><div class="login-title">Acesso ao painel</div><div class="login-sub">Informe seu primeiro e segundo nome para continuar.</div></div></div>',unsafe_allow_html=True)
    with st.form("dashboard_login",clear_on_submit=False):
        name=st.text_input("Primeiro e segundo nome",placeholder="")
        submitted=st.form_submit_button("ENTRAR",use_container_width=True)
    if submitted:
        display,status=authenticate_dashboard_name(cfg,name)
        if status=="ok":
            st.session_state.dashboard_autenticado=True
            st.session_state.dashboard_usuario=display
            token=issue_dashboard_token(st,display)
            if token:
                st.session_state.dashboard_auth_token=token
                st.query_params["auth"]=token
            st.rerun()
        else:
            st.error("Usuário não autorizado. Informe o primeiro e o segundo nome cadastrados.")

def open_seller_dialog(st,x):
    @st.dialog(x["vendedor"],width="large")
    def _dialog():
        status,c,_=performance(x["media"])
        st.markdown(f'<div class="seller-dialog-meta" style="--seller-color:{c}"><span>{html.escape(x.get("equipe","Equipe Interna"))}</span><b>Performance {status}</b></div>',unsafe_allow_html=True)
        st.markdown(seller_kpis_html(x),unsafe_allow_html=True)
        if st.button("FECHAR",key="close_seller_dialog",use_container_width=True):
            st.session_state.pop("seller_detail",None)
            st.rerun()
    _dialog()

def render_management(st,base,current_rows,current_cfg,metadata):
    st.markdown('<div class="section">GESTÃO · ACESSO RESTRITO</div>',unsafe_allow_html=True)
    if not st.session_state.get("gestor_autenticado"):
        password=manager_password(st)
        if not password:st.error("A senha do gestor não foi configurada. Defina GESTOR_SENHA nos Secrets do Streamlit."); return
        supplied=st.text_input("Senha do gestor",type="password")
        if st.button("ENTRAR NA GESTÃO"):
            if hmac.compare_digest(supplied,password):st.session_state.gestor_autenticado=True; st.rerun()
            else:st.error("Senha inválida.")
        return

    st.info("HISTÓRICO DIÁRIO ATIVO: envie somente o relatório do dia anterior. Se uma data já existir, ela será substituída; os demais dias permanecerão acumulados.")
    uploaded=st.file_uploader("IMPORTAR RELATÓRIO DIÁRIO",type=["xlsx","csv"],key="gestao_upload")
    rows=current_rows; cfg=merge_registry(base,current_cfg); source=metadata.get("arquivo","base atual")
    rows,history=ensure_import_history(rows,list(metadata.get("historico_importacoes",[]))); imported_days=[]; pending_import_id=None; pending_import_count=0; pending_seller_count=0
    if uploaded:
        try:
            raw=rows_from_xlsx(uploaded.getvalue()) if uploaded.name.lower().endswith(".xlsx") else rows_from_csv(uploaded.getvalue())
            incoming,_=canonicalize(raw,base)
            pending_import_id=str(uuid.uuid4())
            for row in incoming:row["importacao_id"]=pending_import_id
            if metadata.get("demonstracao"):rows=[]; history=[]
            imported_days=sorted({x["data_venda"] for x in incoming})
            history=mark_replaced_imports(history,imported_days)
            rows,imported_days=merge_daily_history(rows,incoming)
            pending_import_count=len(incoming); pending_seller_count=len({normalize_text(x["vendedor"]) for x in incoming})
            ref=max(imported_days); month_rows=[x for x in rows if x["data_venda"].year==ref.year and x["data_venda"].month==ref.month]
            cfg=prepare_config(merge_registry(base,current_cfg),month_rows,ref.month,ref.year)
            cfg["dia_referencia"]=max(x["data_venda"].day for x in month_rows); source=uploaded.name
        except Exception as exc:st.error(f"O relatório não pôde ser validado: {exc}"); return
    month_rows=[x for x in rows if x["data_venda"].year==cfg["ano"] and x["data_venda"].month==cfg["mes"]]
    dates=[x["data_venda"] for x in month_rows] or [date(cfg["ano"],cfg["mes"],1)]
    unknown=sum(not x.get("classificado",True) for x in cfg["vendedores"]); local=sum(x.get("pertence_franquia",False) for x in cfg["vendedores"])
    cards(st,[("PERÍODO",f"{min(dates):%d/%m/%Y} a {max(dates):%d/%m/%Y}","cyan","Histórico acumulado"),("VENDAS ACUMULADAS",len(month_rows),"green","Competência atual"),("VENDEDORES",len(cfg["vendedores"]),"cyan",f"{local} locais"),("NÃO CLASSIFICADOS",unknown,"yellow","Revisar abaixo")])
    if uploaded and imported_days:st.success("Datas detectadas no novo arquivo: "+", ".join(d.strftime("%d/%m/%Y") for d in imported_days))
    if unknown:st.warning(f"{unknown} vendedores aguardam classificação. Eles ficam em OUTROS CANAIS até você classificá-los/ativá-los corretamente.")

    st.markdown("#### PERÍODO DE REFERÊNCIA")
    month_names=("Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro")
    st.info(f'{month_names[int(cfg["mes"])-1]} / {cfg["ano"]} · cálculos de dias, semanas e projeções continuam usando a lógica atual do sistema.')

    # Metas gerenciais. A meta da empresa mantém a mesma variável já usada nos cálculos;
    # as metas de equipe são apenas novos parâmetros de acompanhamento visual.
    meta_col1,meta_col2,meta_col3=st.columns(3)
    cfg["meta_empresa"]=int(meta_col1.number_input("Meta do mês",min_value=0,value=int(cfg.get("meta_empresa",0)),step=1,key="gestao_meta_empresa"))
    cfg["meta_equipe_interna"]=int(meta_col2.number_input("Meta Equipe Interna",min_value=0,value=int(cfg.get("meta_equipe_interna",0)),step=1,key="gestao_meta_equipe_interna"))
    cfg["meta_equipe_externa"]=int(meta_col3.number_input("Meta Equipe Externa",min_value=0,value=int(cfg.get("meta_equipe_externa",0)),step=1,key="gestao_meta_equipe_externa"))

    sales_by_seller={}
    for row in month_rows:
        key=normalize_text(row.get("vendedor",""))
        if key:sales_by_seller[key]=sales_by_seller.get(key,0)+1

    with st.form("gestao_config"):
        st.markdown("#### GESTÃO DE VENDEDORES")
        st.caption("Os vendedores são identificados automaticamente pela planilha. Aqui você define somente a equipe e se aparecem no Dashboard/Ranking.")
        for i,seller in enumerate(cfg["vendedores"]):
            key=normalize_text(seller.get("vendedor","")); qty=sales_by_seller.get(key,0)
            with st.container(border=True):
                left,mid,right=st.columns([3.0,2.0,1.5])
                left.markdown(f'**{seller["vendedor"]}**')
                left.caption(f'{qty} venda(s) reconhecida(s) no mês')
                current_team=normalized_team(seller.get("equipe"),seller)
                seller["equipe"]=mid.selectbox("Equipe",TEAM_OPTIONS,index=TEAM_OPTIONS.index(current_team),key=f"geq{i}")
                was_active=bool(seller.get("ativo",False))
                seller["ativo"]=right.checkbox("Exibir no Dashboard/Ranking",was_active,key=f"gat{i}")
                if seller["ativo"]:
                    seller["pertence_franquia"]=True
                    if normalize_text(seller.get("categoria")) in {"canal nacional",""}:seller["categoria"]="Vendedor"
                seller["classificado"]=True
        confirmed=st.form_submit_button("SALVAR CONFIGURAÇÕES",use_container_width=True)

    st.markdown("#### RELATÓRIO GERAL DA EQUIPE")
    try:
        management_summary,management_days,management_elapsed,management_official=summarize(rows,cfg)
        apply_team_labels(management_summary,cfg)
        management_team=regular(management_summary)
        general_book=general_report_xlsx_bytes(management_team,management_days)
        validate_xlsx_bytes(general_book,"RELATORIO GERAL")
        st.download_button("BAIXAR RELATÓRIO GERAL DA EQUIPE (EXCEL)",general_book,"Relatorio_Geral_Equipe_Afogados.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,key="gestao_download_relatorio_geral")
    except Exception as exc:
        st.error(f"Não foi possível gerar o Relatório Geral em Excel: {exc}")

    st.markdown("#### ACOMPANHAMENTO SEMANAL")
    st.caption(f'Competência: {int(cfg["mes"]):02d}/{cfg["ano"]} · visão analítica gerencial das mesmas semanas e premiações usadas no ranking.')
    try:
        if 'management_summary' not in locals():
            management_summary,management_days,management_elapsed,management_official=summarize(rows,cfg)
            apply_team_labels(management_summary,cfg)
            management_team=regular(management_summary)
        management_max_weeks=max((len(x.get("semanas",[])) for x in management_team),default=0)
        if management_max_weeks:
            st.markdown(weekly_management_table_html(management_team,management_max_weeks),unsafe_allow_html=True)
            management_sellers=[x["vendedor"] for x in sorted(management_team,key=lambda z:normalize_text(z["vendedor"]))]
            chosen=st.selectbox("Histórico semanal individual",["SELECIONE UM VENDEDOR"]+management_sellers,key="gestao_weekly_seller")
            if chosen!="SELECIONE UM VENDEDOR":
                selected=next((x for x in management_team if x["vendedor"]==chosen),None)
                if selected:st.markdown(weekly_seller_history_html(selected),unsafe_allow_html=True)
        else:st.caption("Sem dados semanais nesta competência.")
    except Exception as exc:
        st.error(f"Não foi possível montar o acompanhamento semanal: {exc}")

    st.markdown("#### PREMIAÇÕES E CENÁRIOS")
    st.caption("Visão restrita à Gestão. Os valores abaixo reutilizam exatamente os mesmos cálculos já existentes no sistema.")
    try:
        if 'management_summary' not in locals():
            management_summary,management_days,management_elapsed,management_official=summarize(rows,cfg)
            apply_team_labels(management_summary,cfg)
            management_team=regular(management_summary)
        management_total=sum(x["vendas"] for x in management_summary)
        management_projection=sum(x["projecao"] for x in management_summary)
        management_projected="maior_ou_igual_1000" if management_projection>=cfg["limite_cenario_maior"] else "abaixo_1000"
        cards(st,[("CENÁRIO ATUAL","≥ 1.000" if management_official=="maior_ou_igual_1000" else "< 1.000","cyan",f"{management_total} vendas"),("CENÁRIO PROJETADO","≥ 1.000" if management_projected=="maior_ou_igual_1000" else "< 1.000","yellow",f"{management_projection} vendas"),("PREMIAÇÃO BASE ATUAL",money(sum(x["base"] for x in management_team)),"cyan",""),("PREMIAÇÃO PROJETADA",money(sum(x["comissao_proj"] for x in management_team)),"yellow","Base projetada"),("BÔNUS NEO PROJ.",money(sum(x["bonus_neo_proj"] for x in management_team)),"green",""),("BÔNUS (SE) 100% ADIM",money(sum(x["bonus_adim_proj"] for x in management_team)),"green",""),("SEMANAIS ACUMULADOS",money(sum(x["premio_total"] for x in management_team)),"cyan",""),("TOTAL VAR. PROJETADO",money(sum(x["total_variavel_proj"] for x in management_team)),"yellow","")])
        st.dataframe([{"Vendedor":x["vendedor"],"Vendas":x["vendas"],"Projeção":x["projecao"],"Mínimo":x["minimo"],"R$/venda":x["taxa"],"Base atual":x["base"],"Premiação projetada":x["comissao_proj"],"Bônus Neo proj.":x["bonus_neo_proj"],"BÔNUS (SE) 100% ADIM":x["bonus_adim_proj"],"Semanais":x["premio_total"],"Total atual":x["total"],"Total var. projetado":x["total_variavel_proj"]} for x in sorted(management_team,key=lambda x:(x["vendas"],x["projecao"]),reverse=True)],use_container_width=True,hide_index=True)
    except Exception as exc:
        st.error(f"Não foi possível montar a área de Premiações: {exc}")

    st.markdown("#### CONFERÊNCIA DA IMPORTAÇÃO")
    audit=[]
    for seller in cfg["vendedores"]:
        qty=sales_by_seller.get(normalize_text(seller.get("vendedor","")),0)
        audit.append({"Vendedor":seller.get("vendedor",""),"Vendas reconhecidas":qty,"Status":"OK"})
    if audit:st.dataframe(audit,use_container_width=True,hide_index=True)
    if confirmed:
        if imported_days and pending_import_id:
            history.append({"importacao_id":pending_import_id,"data_importacao":datetime.now().isoformat(timespec="seconds"),"arquivo":source,"dias":[d.isoformat() for d in imported_days],"registros_arquivo":pending_import_count,"vendedores":pending_seller_count,"usuario":st.session_state.get("dashboard_usuario","") or "","status":"Ativa"}); history=history[-180:]
        save_published(rows,cfg,source,history); st.success("Histórico atualizado e dashboard publicado."); st.rerun()
    if st.session_state.pop("import_delete_flash",None):st.success(st.session_state.pop("import_delete_flash_text","Importação excluída com sucesso."))
    if history:
        st.markdown("#### HISTÓRICO DE IMPORTAÇÕES")
        st.caption("Excluir uma importação remove efetivamente as vendas daquele lote e recalcula o dashboard. Importações substituídas permanecem apenas para auditoria.")
        for h in reversed(history[-30:]):
            status=h.get("status","Ativa"); import_id=h.get("importacao_id",""); days=", ".join(h.get("dias",[])); regs=int(h.get("registros_arquivo",0) or 0); sellers_count=int(h.get("vendedores",0) or 0)
            with st.container(border=True):
                a,b,c,d=st.columns([2.1,2.2,1.1,1.1])
                a.markdown(f'**{h.get("arquivo","Importação")}**')
                a.caption(h.get("data_importacao","").replace("T"," "))
                b.markdown(f'**Período:** {days or "—"}')
                b.caption(f'Usuário: {h.get("usuario","") or "—"}')
                c.metric("Vendas",regs); d.metric("Vendedores",sellers_count)
                st.caption(f'Status: **{status}**')
                if status=="Ativa" and import_id:
                    if st.button("EXCLUIR IMPORTAÇÃO",key=f'del_import_{import_id}',type="secondary"):
                        st.session_state["confirm_delete_import_id"]=import_id; st.rerun()
                elif status=="Substituída":
                    st.caption("Esta importação foi substituída por uma importação posterior do mesmo dia e não possui vendas ativas no dashboard.")
                else:
                    st.caption("Importação excluída. Nenhuma venda deste lote permanece ativa.")

        confirm_id=st.session_state.get("confirm_delete_import_id")
        target=next((h for h in history if h.get("importacao_id")==confirm_id and h.get("status")=="Ativa"),None) if confirm_id else None
        if target:
            affected=[x for x in rows if x.get("importacao_id")==confirm_id]
            affected_sellers=len({normalize_text(x.get("vendedor")) for x in affected})
            @st.dialog("Confirmar exclusão da importação")
            def _confirm_delete_import():
                st.warning(f'Tem certeza de que deseja excluir esta importação? Serão removidas {len(affected)} vendas referentes a {", ".join(target.get("dias",[])) or "este lote"} e todos os indicadores serão recalculados.')
                a,b=st.columns(2)
                if a.button("CANCELAR",use_container_width=True):
                    st.session_state.pop("confirm_delete_import_id",None); st.rerun()
                if b.button("CONFIRMAR EXCLUSÃO",type="primary",use_container_width=True):
                    new_rows=[x for x in rows if x.get("importacao_id")!=confirm_id]
                    target["status"]="Excluída"; target["data_exclusao"]=datetime.now().isoformat(timespec="seconds"); target["excluido_por"]=st.session_state.get("dashboard_usuario","") or ""
                    remaining_month=[x for x in new_rows if x["data_venda"].year==cfg["ano"] and x["data_venda"].month==cfg["mes"]]
                    if remaining_month:cfg["dia_referencia"]=max(x["data_venda"].day for x in remaining_month)
                    else:cfg["dia_referencia"]=1
                    save_published(new_rows,cfg,metadata.get("arquivo","base atual"),history)
                    st.session_state.pop("confirm_delete_import_id",None)
                    st.session_state["import_delete_flash"]=True
                    st.session_state["import_delete_flash_text"]=f'IMPORTAÇÃO EXCLUÍDA COM SUCESSO · {len(affected)} vendas removidas · {affected_sellers} vendedores impactados · dashboard recalculado.'
                    st.rerun()
            _confirm_delete_import()
    if st.button("SAIR DA GESTÃO"):st.session_state.gestor_autenticado=False; st.rerun()

def render_app():
    import streamlit as st
    st.set_page_config(page_title="Painel Comercial — Afogados",page_icon="📊",layout="wide",initial_sidebar_state="collapsed")
    st.markdown(CSS,unsafe_allow_html=True)
    base=json.loads((ROOT/"config.json").read_text(encoding="utf-8"))
    global BASE_SELLER_DEFAULTS
    BASE_SELLER_DEFAULTS=copy.deepcopy(base.get("vendedores",[]))
    try:rows,cfg,metadata=load_published(base)
    except Exception as exc:st.error(f"A base publicada não pôde ser carregada: {exc}");return
    incoming_token=st.query_params.get("auth")
    if not st.session_state.get("dashboard_autenticado",False) and incoming_token:
        restored_user=validate_dashboard_token(st,cfg,incoming_token)
        if restored_user:
            st.session_state.dashboard_autenticado=True
            st.session_state.dashboard_usuario=restored_user
            st.session_state.dashboard_auth_token=incoming_token
    if not st.session_state.get("dashboard_autenticado",False):
        render_login(st,cfg)
        return
    areas=["VISÃO GERAL","SEMANAL"]
    if st.session_state.get("area") not in areas+["GESTÃO"]:st.session_state.area="VISÃO GERAL"
    action=st.query_params.get("action")
    if action=="logout":
        for key in ("dashboard_autenticado","dashboard_usuario","dashboard_auth_token","seller_detail","gestor_autenticado","login_duplicate_first"):
            st.session_state.pop(key,None)
        st.session_state.area="VISÃO GERAL"
        st.query_params.clear()
        st.rerun()
    if action=="management":
        st.session_state.area="GESTÃO"
        try:del st.query_params["action"]
        except KeyError:pass
    user_name=html.escape(str(st.session_state.get("dashboard_usuario") or "Usuário autenticado"))
    auth_token=st.session_state.get("dashboard_auth_token") or incoming_token or ""
    management_href=f'?auth={html.escape(str(auth_token),quote=True)}&action=management' if auth_token else '?action=management'
    logout_href=f'?auth={html.escape(str(auth_token),quote=True)}&action=logout' if auth_token else '?action=logout'
    st.markdown(
        '<div class="bi-topbar bi-topbar-nav integrated-header">'
        '<div class="bi-brand"><h1>PAINEL COMERCIAL — AFOGADOS</h1><p>Visão executiva de produção, performance, histórico e remuneração variável</p></div>'
        f'<div class="header-account"><span class="header-user">{user_name}</span><div class="header-actions"><a href="{management_href}" target="_self">GESTÃO</a><a href="{logout_href}" target="_self">SAIR</a></div></div>'
        '</div>',unsafe_allow_html=True
    )
    if st.session_state.area=="GESTÃO":
        render_management(st,base,rows,cfg,metadata)
        return
    selected_area=st.segmented_control("Navegação",areas,default=st.session_state.area,key="top_nav_area",label_visibility="collapsed")
    if selected_area and selected_area != st.session_state.area:
        st.session_state.area=selected_area
        st.rerun()
    area=st.session_state.area
    try:summary,all_days,elapsed,official=summarize(rows,cfg); apply_team_labels(summary,cfg)
    except Exception as exc:st.error(f"Falha ao processar relatório: {exc}");return
    data_until=max((x["data_venda"] for x in rows if x["data_venda"].year==cfg["ano"] and x["data_venda"].month==cfg["mes"]),default=date(cfg["ano"],cfg["mes"],1)); updated=datetime.fromisoformat(metadata["atualizado_em"])
    st.caption(f"Última atualização: {updated:%d/%m/%Y às %H:%M}  •  Dados acumulados até: {data_until:%d/%m/%Y}  •  Competência: {cfg['mes']:02d}/{cfg['ano']}")
    team=regular(summary); total=sum(x["vendas"] for x in summary); projection=sum(x["projecao"] for x in summary); neo=sum(x["neo"] for x in summary); color=lambda x:performance(x["media"])[1]
    # Abrir detalhamento somente após clique explícito no vendedor.
    # Limpa estado legado para filtros/equipe não reabrirem o último popup.
    st.session_state.pop("seller_detail",None)
    requested_seller=st.query_params.get("seller")
    if requested_seller:
        try:del st.query_params["seller"]
        except KeyError:pass
        detail=next((x for x in team if normalize_text(x["vendedor"])==normalize_text(requested_seller)),None)
        if detail:open_seller_dialog(st,detail)
    if area=="VISÃO GERAL":
        st.markdown(executive_kpis_html(cfg,total,projection,neo,team),unsafe_allow_html=True)
        st.markdown('<div class="section">Distribuição de performance</div>',unsafe_allow_html=True); counts={k:0 for k in ("Azul","Verde","Amarelo","Vermelho")}
        for x in team:counts[performance(x["media"])[0]]+=1
        st.markdown(performance_summary_html(counts),unsafe_allow_html=True)
        st.markdown('<div class="section">Ranking da equipe</div>',unsafe_allow_html=True)
        team_filter=st.selectbox("Filtrar ranking por equipe",("TODAS AS EQUIPES",)+TEAM_OPTIONS,key="ranking_team_filter",label_visibility="collapsed")
        filtered_team=team if team_filter=="TODAS AS EQUIPES" else [x for x in team if x.get("equipe")==team_filter]
        ranking=sorted(filtered_team,key=lambda x:(x["vendas"],x["projecao"]),reverse=True); st.markdown(ranking_html(ranking,auth_token),unsafe_allow_html=True)
        st.markdown('<div class="section">Produção por canal</div>',unsafe_allow_html=True); channels={name:0 for name in ("VENDEDORES FRANQUIA","CANAL NACIONAL")}
        for item in summary:
            name=channel_name(item)
            if name in channels:channels[name]+=item["vendas"]
        st.markdown(production_channel_dashboard_html(channels,total,summary,cfg,data_until,updated),unsafe_allow_html=True)
        render_general_report(st,team,rows,cfg,summary,all_days,elapsed,official,color)
    elif area=="SEMANAL":
        st.markdown('<div class="section">Acompanhamento semanal</div>',unsafe_allow_html=True)
        if not team:st.warning("Nenhum vendedor local ativo.");return
        max_weeks=max(len(x.get("semanas",[])) for x in team)
        current_index=weekly_current_index(cfg,max_weeks)
        week_labels=weekly_week_labels(cfg,max_weeks,current_index)
        selector_key=f'weekly_week_selector_gamified_{cfg["ano"]}_{cfg["mes"]}_{current_index}'
        selected_week=st.segmented_control("Semana",week_labels,default=week_labels[current_index],key=selector_key,label_visibility="collapsed") or week_labels[current_index]
        week_index=week_labels.index(selected_week)
        is_current=(week_index==current_index)
        if week_index<current_index:
            st.caption(f'Semana {week_index+1} encerrada · histórico final')
        elif is_current:
            st.caption(f'Semana {week_index+1} atual · ranking em andamento')
        else:
            st.caption(f'Semana {week_index+1} · período futuro da competência')
        st.markdown('<div class="section">Ranking da semana</div>',unsafe_allow_html=True)
        st.markdown(weekly_rank_gamified_html(team,week_index,cfg,is_current),unsafe_allow_html=True)

if __name__=="__main__":render_app()
